import json
from datetime import timedelta
from urllib.error import URLError
from urllib.request import urlopen
from io import BytesIO

import pandas as pd
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.utils.translation import gettext as _
from django.utils import timezone
from django.urls import reverse
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Avg

from .forms import (
    RiskSimulationForm,
    SoilMetricsForm,
    ClimateMetricsForm,
    DroughtIndicesForm,
    AgriculturalMetricsForm,
    RemoteSensingMetricsForm,
    HydrologyMetricsForm,
    ExcelImportForm,
    BulkMetricsImportForm,
    UserCreateForm,
)
from .models import (
    CropType,
    EnvironmentalSnapshot,
    IrrigationMethod,
    ObservationYear,
    Region,
    RiskAssessment,
    DroughtPrediction,
    Sensor,
    SoilMetrics,
    ClimateMetrics,
    DroughtIndices,
    AgriculturalMetrics,
    RemoteSensingMetrics,
    HydrologyMetrics,
    DataImportLog,
    UserProfile,
)
from .data_ingestion_utils import DataImporter, DataIngestionError, create_reference_data


def _seed_reference_data():
    needs_seed = Region.objects.filter(name__in=["Tunisia", "Morocco", "Algeria", "Bizerte"]).count() < 4
    if needs_seed:
        CropType.objects.filter(name="Ble (Durum)").update(name="Blé (Durum)")
        IrrigationMethod.objects.filter(name="Irrigation au goutte-a-goutte").update(
            name="Irrigation au goutte-à-goutte"
        )
        IrrigationMethod.objects.filter(name="Alimente par la pluie").update(
            name="Alimenté par la pluie"
        )

        for region in ["Tunisia", "Morocco", "Algeria", "Bizerte"]:
            Region.objects.get_or_create(name=region)
        for year in ["2026", "2025", "2024", "2023", "2022", "2020 - 2024", "2015 - 2019"]:
            ObservationYear.objects.get_or_create(label=year)
        for crop in ["Blé (Durum)", "Olives", "Tomates", "Orge"]:
            CropType.objects.get_or_create(name=crop)
        for irrigation in [
            "Irrigation au goutte-à-goutte",
            "Asperseurs",
            "Alimenté par la pluie",
        ]:
            IrrigationMethod.objects.get_or_create(name=irrigation)

        region = Region.objects.get(name="Tunisia")
        year = ObservationYear.objects.get(label="2026")
        EnvironmentalSnapshot.objects.get_or_create(
            region=region,
            year=year,
            defaults={
                "wind_speed_kmh": 24,
                "wind_gust_kmh": 32,
                "wind_direction": "NE",
                "rainfall_mm": 18.5,
                "rainfall_delta_percent": -12,
                "ph_level": 7.2,
                "npk_index": "Med-High",
                "temperature_c": 28,
                "humidity_percent": 64,
            },
        )

    # Always clean up duplicate (Current) year labels and ensure current year exists
    ObservationYear.objects.filter(label__endswith="(Current)").delete()
    ObservationYear.objects.get_or_create(label=str(timezone.now().year))

    bizerte = Region.objects.filter(name="Bizerte").first()
    if bizerte:
        from decimal import Decimal
        changed = False
        if bizerte.latitude is None:
            bizerte.latitude = Decimal("37.27")
            changed = True
        if bizerte.longitude is None:
            bizerte.longitude = Decimal("9.87")
            changed = True
        if bizerte.country is None:
            bizerte.country = "Tunisia"
            changed = True
        if bizerte.area_km2 is None:
            bizerte.area_km2 = Decimal("3680")
            changed = True
        if changed:
            bizerte.save()


# ============== RBAC HELPERS ==============


def _is_admin_user(user):
    """Check if user is superadmin (full access)."""
    if not user.is_authenticated:
        return False
    return getattr(user, 'profile', None) and user.profile.role == 'superadmin'


def _is_subadmin_user(user):
    """Check if user is subadmin (region-restricted access)."""
    if not user.is_authenticated:
        return False
    return getattr(user, 'profile', None) and user.profile.role == 'subadmin'


def _is_viewer_user(user):
    """Check if user is viewer (read-only access)."""
    if not user.is_authenticated:
        return False
    return getattr(user, 'profile', None) and user.profile.role == 'viewer'


def _deny_admin_access(request):
    """Redirect non-superadmin users away from admin pages."""
    messages.error(request, _('You do not have permission to access this page.'))
    return redirect('dashboard:dashboard')


def _deny_data_access(request):
    """Redirect users without data mutation permission."""
    messages.error(request, _('You do not have permission to modify data.'))
    return redirect('dashboard:dashboard')


def _user_can_access_region(user, region_id):
    """Check if a subadmin can access the given region."""
    if _is_admin_user(user):
        return True
    if _is_subadmin_user(user):
        return str(user.profile.region_id) == str(region_id)
    return False


def _check_data_permission(request, region_id=None):
    """Check if user can mutate data.
    - Viewer: always denied
    - Subadmin: allowed only if region_id matches their assigned region
    - Superadmin: always allowed
    """
    if not request.user.is_authenticated:
        return False
    profile = getattr(request.user, 'profile', None)
    if not profile:
        return False
    if profile.is_viewer():
        return False
    if profile.is_superadmin():
        return True
    if profile.is_subadmin():
        if region_id is not None:
            return str(profile.region_id) == str(region_id)
        return True
    return False


# ============== COORDINATE VALIDATION ==============

import math


def _haversine_distance(lat1, lon1, lat2, lon2):
    """Calculate distance in km between two lat/lng pairs (Haversine formula)."""
    R = 6371
    dlat = math.radians(float(lat2) - float(lat1))
    dlon = math.radians(float(lon2) - float(lon1))
    a = math.sin(dlat / 2) ** 2 + math.cos(math.radians(float(lat1))) * math.cos(math.radians(float(lat2))) * math.sin(dlon / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c


def _validate_coordinates_for_region(region, latitude, longitude):
    """Check if lat/lng falls within the region's radius from its center.
    Returns (is_valid: bool, error_message: str or None).
    """
    if not region.latitude or not region.longitude:
        return True, None
    if latitude is None or longitude is None:
        return False, _('Latitude and longitude are required.')
    try:
        lat = float(latitude)
        lng = float(longitude)
    except (TypeError, ValueError):
        return False, _('Invalid latitude or longitude values.')
    if not (-90 <= lat <= 90) or not (-180 <= lng <= 180):
        return False, _('Latitude must be between -90 and 90, longitude between -180 and 180.')
    distance = _haversine_distance(region.latitude, region.longitude, lat, lng)
    radius = float(region.radius_km) if region.radius_km else 100.0
    if distance > radius:
        return False, _('Coordinates are {:.0f} km from the center of {}, which exceeds the {:.0f} km limit.').format(distance, region.name, radius)
    return True, None


# ============== METRIC BROWSER ==============


def _metric_browser_value(value):
    if value is None or value == "":
        return "—"
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


METRIC_BROWSER_CONFIGS = {
    "soil": {
        "title": "Soil Metrics",
        "model": SoilMetrics,
        "form": SoilMetricsForm,
        "select_related": ("region", "year"),
        "add_url": "dashboard:add_soil_metrics",
        "export_type": "soil",
    },
    "climate": {
        "title": "Climate Metrics",
        "model": ClimateMetrics,
        "form": ClimateMetricsForm,
        "select_related": ("region", "year"),
        "add_url": "dashboard:add_climate_metrics",
        "export_type": "climate",
    },
    "drought": {
        "title": "Drought Indices",
        "model": DroughtIndices,
        "form": DroughtIndicesForm,
        "select_related": ("region", "year"),
        "add_url": "dashboard:add_drought_indices",
        "export_type": "drought",
    },
    "agricultural": {
        "title": "Agricultural Metrics",
        "model": AgriculturalMetrics,
        "form": AgriculturalMetricsForm,
        "select_related": ("region", "year", "crop", "irrigation_method"),
        "add_url": "dashboard:add_agricultural_metrics",
        "export_type": "agricultural",
    },
    "remote_sensing": {
        "title": "Remote Sensing Metrics",
        "model": RemoteSensingMetrics,
        "form": RemoteSensingMetricsForm,
        "select_related": ("region", "year"),
        "add_url": "dashboard:add_remote_sensing_metrics",
        "export_type": "remote_sensing",
    },
    "hydrology": {
        "title": "Hydrology Metrics",
        "model": HydrologyMetrics,
        "form": HydrologyMetricsForm,
        "select_related": ("region", "year"),
        "add_url": "dashboard:add_hydrology_metrics",
        "export_type": "hydrology",
    },
}


def _build_browser_row(obj):
    values = []
    search_bits = []
    for field in obj._meta.fields:
        display = _metric_browser_value(getattr(obj, field.name))
        values.append(display)
        search_bits.append(display.lower())

    return {
        "pk": obj.pk,
        "values": values,
        "search_text": " ".join(search_bits).lower(),
        "filters": {
            "region_id": getattr(obj, "region_id", ""),
            "year_id": getattr(obj, "year_id", ""),
            "crop_id": getattr(obj, "crop_id", ""),
        },
    }


def _build_browser_section(key, config):
    total_count = config["model"].objects.count()
    queryset = config["model"].objects.select_related(*config["select_related"]).order_by("-id")[:25]
    rows = [_build_browser_row(obj) for obj in queryset]
    return {
        "key": key,
        "title": config["title"],
        "model_name": config["model"].__name__,
        "headers": [field.verbose_name.title() for field in config["model"]._meta.fields],
        "rows": rows,
        "count": total_count,
        "display_count": len(rows),
        "add_url": reverse(config["add_url"]),
        "export_url": reverse("dashboard:export_metrics", args=[config["export_type"]]),
    }


def metrics_browser(request):
    """Browse all metric tables with edit/delete actions and client-side filters."""
    sections = [_build_browser_section(key, config) for key, config in METRIC_BROWSER_CONFIGS.items()]
    context = {
        "title": "View All Metrics",
        "sections": sections,
        "regions": Region.objects.order_by("name"),
        "years": ObservationYear.objects.order_by("-label"),
        "crops": CropType.objects.order_by("name"),
        "section_count": len(sections),
        "total_rows": sum(section["count"] for section in sections),
    }
    return render(request, "dashboard/metrics_browser.html", context)


def edit_metric_entry(request, metric_type, pk):
    """Edit a single metric entry."""
    config = METRIC_BROWSER_CONFIGS.get(metric_type)
    if not config:
        messages.error(request, _("Unknown metric type."))
        return redirect("dashboard:metrics_browser")

    instance = get_object_or_404(config["model"], pk=pk)
    region_id = getattr(instance, 'region_id', None)
    if not _check_data_permission(request, region_id):
        return _deny_data_access(request)

    form_class = config["form"]

    if request.method == "POST":
        form = form_class(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, _("%s updated successfully!") % config["title"])
            return redirect(reverse("dashboard:metrics_browser") + f"#{metric_type}")
    else:
        form = form_class(instance=instance)

    context = {
        "title": f"Edit {config['title']}",
        "form": form,
        "metric_type": metric_type,
        "back_url": reverse("dashboard:metrics_browser") + f"#{metric_type}",
        "section_title": config["title"],
        "record_label": str(instance),
    }
    return render(request, "dashboard/metrics_form.html", context)


@require_http_methods(["POST"])
def delete_metric_entry(request, metric_type, pk):
    """Delete a single metric entry."""
    config = METRIC_BROWSER_CONFIGS.get(metric_type)
    if not config:
        messages.error(request, _("Unknown metric type."))
        return redirect("dashboard:metrics_browser")

    instance = get_object_or_404(config["model"], pk=pk)
    region_id = getattr(instance, 'region_id', None)
    if not _check_data_permission(request, region_id):
        return _deny_data_access(request)

    instance.delete()
    messages.success(request, _("%s deleted successfully!") % config["title"])
    return redirect(reverse("dashboard:metrics_browser") + f"#{metric_type}")


@require_http_methods(["POST"])
def delete_metric_table(request, metric_type):
    """Delete all entries for a given metric type. Region-restricted for subadmins."""
    config = METRIC_BROWSER_CONFIGS.get(metric_type)
    if not config:
        messages.error(request, _("Unknown metric type."))
        return redirect("dashboard:metrics_browser")

    if _is_viewer_user(request.user):
        return _deny_data_access(request)

    model = config["model"]
    if _is_admin_user(request.user):
        model.objects.all().delete()
    elif _is_subadmin_user(request.user):
        region_id = request.user.profile.region_id
        if region_id:
            model.objects.filter(region_id=region_id).delete()
        else:
            messages.error(request, _("No region assigned to your account."))
            return redirect("dashboard:metrics_browser")
    else:
        return _deny_data_access(request)

    messages.success(request, _("All %s deleted successfully!") % config["title"])
    return redirect(reverse("dashboard:metrics_browser") + f"#{metric_type}")


@require_http_methods(["POST"])
def delete_all_metrics(request):
    """Delete all metric data across all tables. Superadmin — full; Subadmin — own region only."""
    if _is_viewer_user(request.user):
        return _deny_data_access(request)

    if _is_admin_user(request.user):
        for config in METRIC_BROWSER_CONFIGS.values():
            config["model"].objects.all().delete()
        messages.success(request, _("All metric data deleted successfully!"))
    elif _is_subadmin_user(request.user):
        region_id = request.user.profile.region_id
        if not region_id:
            messages.error(request, _("No region assigned to your account."))
            return redirect("dashboard:metrics_browser")
        for config in METRIC_BROWSER_CONFIGS.values():
            config["model"].objects.filter(region_id=region_id).delete()
        messages.success(request, _("All data for your region deleted successfully!"))
    else:
        return _deny_data_access(request)

    return redirect("dashboard:metrics_browser")


def _calculate_risk(crop_name, irrigation_name):
    crop_lower = crop_name.lower()
    irr_lower = irrigation_name.lower()
    wheat_like = "blé" in crop_lower or crop_lower.startswith("ble") or "durum" in crop_lower or "wheat" in crop_lower
    drip_like = "goutte" in irr_lower or "drip" in irr_lower
    if wheat_like and drip_like:
        return (
            _("High risk"),
            _(
                "Current soil salinity levels in sector 4 are incompatible with durum wheat under drip irrigation. "
                "Consider switching to highly salt-tolerant crops or scheduling intensive leaching protocols before sowing."
            ),
        )
    if "olive" in crop_lower or "oliv" in crop_lower:
        return (
            _("Moderate risk"),
            _("The crop is generally suitable. Monitor soil moisture and optimize water application."),
        )
    return (
        _("Low risk"),
        _("Current indicators are favorable. Continue weekly monitoring."),
    )


def _degree_to_compass(degree):
    directions = ["N", "NE", "E", "SE", "S", "SO", "O", "NO"]
    idx = int((degree % 360) / 45 + 0.5) % 8
    return directions[idx]


def _region_map_config():
    return {
        "tunisia": {
            "center": [37.13, 9.67],
            "zoom": 8,
            "label": _("Ichkeul National Park, Tunisia"),
        },
        "morocco": {
            "center": [34.02, -6.84],
            "zoom": 8,
            "label": _("Rabat area, Morocco"),
        },
        "algeria": {
            "center": [36.75, 3.06],
            "zoom": 8,
            "label": _("Algiers area, Algeria"),
        },
    }


def _lakes_by_region():
    return {
        "tunisia": [
            {"key": "ichkeul", "label": _("Lake Ichkeul"), "center": [37.16, 9.66], "zoom": 11},
            {"key": "ghezala", "label": _("Lake Ghezala"), "center": [37.03, 9.34], "zoom": 11},
            {"key": "bizerte", "label": _("Lake of Bizerte"), "center": [37.23, 9.89], "zoom": 11},
        ],
        "morocco": [
            {"key": "dayet-aoua", "label": _("Dayet Aoua"), "center": [33.53, -5.02], "zoom": 11},
            {"key": "afennourir", "label": _("Lake Afennourir"), "center": [33.28, -5.35], "zoom": 11},
            {"key": "bin-el-ouidane", "label": _("Bin el Ouidane"), "center": [32.11, -6.45], "zoom": 11},
        ],
        "algeria": [
            {"key": "oubeira", "label": _("Lake Oubeira"), "center": [36.86, 8.44], "zoom": 11},
            {"key": "fetzara", "label": _("Lake Fetzara"), "center": [36.82, 7.53], "zoom": 11},
            {"key": "melah", "label": _("Lake Mellah"), "center": [36.90, 8.33], "zoom": 11},
        ],
    }


def _dashboard_js_i18n():
    return {
        "map_centered_on": _("Map centered on __PLACE__."),
        "drawings_cleared": _("All drawings have been removed."),
        "select_mode": _("Selection mode: click a drawn area to see its details."),
        "draw_mode": _("Draw mode: trace a polygon to measure the area."),
        "measure_mode": _("Ruler mode: click two points on the map to measure distance."),
        "ping_mode": _("Ping mode: click a point for exact live metrics."),
        "ruler_cleared": _("Ruler measurement cleared."),
        "ping_no_live": _("Ping saved, but live metrics are unavailable for now."),
        "first_point_done": _("First point recorded. Click the second point."),
        "distance_line": _("Distance: __M__ m (__KM__ km)."),
        "polygon_line": _(
            "Area: __M2__ m² (__HA__ ha, __KM2__ km²) | Perimeter: __PM__ m | Center: __LAT__, __LNG__"
        ),
        "ping_line": _(
            "Ping: __LAT__, __LNG__ | Wind: __WS__ km/h | Temp: __T__°C | Humidity: __H__% | Rain: __R__ mm"
        ),
        "sensor_drop_mode": _("Sensor drop mode: click the map to place a sensor pin."),
        "sensor_coords_filled": _("Coordinates filled. Complete the form and submit."),
    }


def live_metrics_api(request):
    try:
        lat = float(request.GET.get("lat", "37.16"))
        lon = float(request.GET.get("lon", "9.66"))
    except ValueError:
        return JsonResponse({"error": "Invalid coordinates"}, status=400)

    api_url = (
        "https://api.open-meteo.com/v1/forecast"
        f"?latitude={lat}&longitude={lon}"
        "&current=temperature_2m,relative_humidity_2m,precipitation,wind_speed_10m,wind_gusts_10m,wind_direction_10m"
        "&timezone=auto"
    )

    try:
        with urlopen(api_url, timeout=10) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except URLError:
        return JsonResponse({"error": "Unable to fetch live metrics"}, status=502)

    current = payload.get("current", {})
    wind_speed = float(current.get("wind_speed_10m", 0))
    wind_gust = float(current.get("wind_gusts_10m", wind_speed))
    wind_deg = float(current.get("wind_direction_10m", 0))
    precipitation = float(current.get("precipitation", 0))
    temperature = float(current.get("temperature_2m", 0))
    humidity = int(current.get("relative_humidity_2m", 0))

    delta = 0
    avg_rainfall = ClimateMetrics.objects.filter(
        rainfall_mm__isnull=False
    ).aggregate(avg=Avg("rainfall_mm"))["avg"]
    if avg_rainfall is not None and avg_rainfall > 0:
        delta = round(((precipitation - float(avg_rainfall)) / float(avg_rainfall)) * 100)

    return JsonResponse(
        {
            "wind_speed_kmh": round(wind_speed, 1),
            "wind_gust_kmh": round(wind_gust, 1),
            "wind_direction": _degree_to_compass(wind_deg),
            "rainfall_mm": round(precipitation, 1),
            "rainfall_delta_percent": delta,
            "temperature_c": round(temperature, 1),
            "humidity_percent": humidity,
            "source": "open-meteo",
        }
    )


def _svg_points(values, vmin=None, vmax=None, svg_x_start=25, svg_x_end=295, svg_y_top=10, svg_y_bottom=130):
    n = len(values)
    if n == 0:
        return ""
    if n == 1:
        x = (svg_x_start + svg_x_end) / 2
        return f"{x:.1f},{((svg_y_top + svg_y_bottom) / 2):.1f}"
    lo = vmin if vmin is not None else min(values)
    hi = vmax if vmax is not None else max(values)
    if hi == lo:
        hi = lo + 1
    pts = []
    for i, v in enumerate(values):
        x = svg_x_start + (i / (n - 1)) * (svg_x_end - svg_x_start)
        norm = (v - lo) / (hi - lo)
        y = svg_y_bottom - norm * (svg_y_bottom - svg_y_top)
        pts.append(f"{x:.1f},{y:.1f}")
    return " ".join(pts)


def _month_labels(queryset=None, limit=12):
    if queryset is None:
        return ["Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
                "Jan", "Feb", "Mar", "Apr", "May"]
    ms = []
    seen = set()
    for obj in queryset:
        if obj.measurement_date:
            m = obj.measurement_date.strftime("%b")
            if m not in seen:
                ms.append(m)
                seen.add(m)
            if len(ms) >= limit:
                break
    ms.reverse()
    while len(ms) < limit:
        ms.insert(0, "")
    return ms[:limit]


def _chart_series(queryset, field, limit=12):
    vals = []
    for obj in queryset:
        v = getattr(obj, field)
        if v is not None:
            vals.append(float(v))
        if len(vals) >= limit:
            break
    vals.reverse()  # newest-first → chronological order
    return vals


def _chart_months(queryset, field, limit=12):
    """Return exactly `limit` month labels aligned 1:1 with _chart_series()."""
    labels = []
    for obj in queryset:
        v = getattr(obj, field, None)
        if v is not None:
            labels.append(obj.measurement_date.strftime("%b") if obj.measurement_date else "")
        if len(labels) >= limit:
            break
    labels.reverse()
    while len(labels) < limit:
        labels.insert(0, "")
    return labels[:limit]


def _normalize(values):
    """Scale values to 0-100% based on data min/max."""
    if not values:
        return values
    lo, hi = min(values), max(values)
    if hi == lo:
        return [50.0] * len(values)
    return [(v - lo) / (hi - lo) * 100 for v in values]


def _year_series(queryset, field, n=12):
    """Extract n evenly-spaced values from a queryset (newest-first) for yearly view."""
    all_vals = []
    for obj in queryset:
        v = getattr(obj, field)
        if v is not None:
            all_vals.append(float(v))
    if not all_vals:
        return []
    all_vals.reverse()  # chronological
    if len(all_vals) <= n:
        return all_vals
    step = (len(all_vals) - 1) / n
    return [all_vals[round(i * step)] for i in range(n)]


def _year_months(queryset, field, n=12):
    """Month labels aligned with _year_series()."""
    all_dates = []
    for obj in queryset:
        v = getattr(obj, field, None)
        if v is not None:
            all_dates.append(obj.measurement_date)
    if not all_dates:
        return [""] * n
    all_dates.reverse()
    if len(all_dates) <= n:
        labels = [d.strftime("%b") if d else "" for d in all_dates]
        while len(labels) < n:
            labels.insert(0, "")
        return labels[:n]
    step = (len(all_dates) - 1) / n
    labels = []
    for i in range(n):
        d = all_dates[round(i * step)]
        labels.append(d.strftime("%b") if d else "")
    return labels


def dashboard_view(request):
    _seed_reference_data()
    result = None
    form = RiskSimulationForm(request.POST or None)
    now = timezone.now()
    year_ago = now - timedelta(days=365)

    if request.method == "POST" and form.is_valid():
        region = form.cleaned_data["region"]
        if not _check_data_permission(request, region.id if region else None):
            messages.error(request, _('You do not have permission to submit risk assessments.'))
            return redirect('dashboard:dashboard')
        year = form.cleaned_data["year"]
        crop = form.cleaned_data["crop"]
        irrigation = form.cleaned_data["irrigation"]
        risk_level, recommendation = _calculate_risk(crop.name, irrigation.name)
        result = {
            "risk_level": risk_level,
            "recommendation": recommendation,
        }
        RiskAssessment.objects.create(
            region=region,
            year=year,
            crop=crop,
            irrigation=irrigation,
            risk_level=risk_level,
            recommendation=recommendation,
            created_by=request.user if request.user.is_authenticated else None,
        )
        snapshot = EnvironmentalSnapshot.objects.filter(region=region, year=year).first()
        chart_region = region
        chart_year = year
    else:
        default_region = Region.objects.filter(name="Bizerte").first() or Region.objects.first()
        default_year = ObservationYear.objects.filter(label=str(timezone.now().year)).first()
        if not default_year:
            default_year = ObservationYear.objects.order_by('-label').first()
        snapshot = EnvironmentalSnapshot.objects.filter(
            region=default_region, year=default_year
        ).first()
        chart_region = default_region
        chart_year = default_year

    # ---- Live chart data ----
    months = _month_labels()

    # Extract year number from chart_year for yearly filtering
    try:
        year_num = int(chart_year.label.split()[0])
    except (ValueError, AttributeError):
        year_num = None

    if year_num:
        climate_qs = ClimateMetrics.objects.filter(
            region=chart_region, measurement_date__year=year_num
        ).order_by("-measurement_date")
        drought_qs = DroughtIndices.objects.filter(
            region=chart_region, measurement_date__year=year_num
        ).order_by("-measurement_date")
        rs_qs = RemoteSensingMetrics.objects.filter(
            region=chart_region, measurement_date__year=year_num
        ).order_by("-measurement_date")

        climate_rainfall = _year_series(climate_qs, "rainfall_mm")
        climate_temp = _year_series(climate_qs, "temperature_mean_c")
        drought_spi = _year_series(drought_qs, "spi_3month")
        drought_spei = _year_series(drought_qs, "spei_3month")
        ndvi_values = _year_series(rs_qs, "ndvi")

        chart_months = _year_months(climate_qs, "rainfall_mm")
    else:
        climate_qs = ClimateMetrics.objects.filter(
            region=chart_region, measurement_date__gte=year_ago
        ).order_by("-measurement_date")
        drought_qs = DroughtIndices.objects.filter(
            region=chart_region, measurement_date__gte=year_ago
        ).order_by("-measurement_date")
        rs_qs = RemoteSensingMetrics.objects.filter(
            region=chart_region, measurement_date__gte=year_ago
        ).order_by("-measurement_date")

        climate_rainfall = _chart_series(climate_qs, "rainfall_mm")
        climate_temp = _chart_series(climate_qs, "temperature_mean_c")
        drought_spi = _chart_series(drought_qs, "spi_3month")
        drought_spei = _chart_series(drought_qs, "spei_3month")
        ndvi_values = _chart_series(rs_qs, "ndvi")

        chart_months = _chart_months(climate_qs, "rainfall_mm")

    if not any(chart_months):
        chart_months = months

    # Normalized (0-100%) versions for the overlaid climate chart
    climate_rainfall_norm = _normalize(climate_rainfall)
    climate_temp_norm = _normalize(climate_temp)
    drought_spi_norm = _normalize(drought_spi)
    ndvi_values_norm = _normalize(ndvi_values)

    # Fixed display ranges for standalone drought/NDVI charts
    spi_min, spi_max = -3, 3
    ndvi_min, ndvi_max = 0, 1

    def _labels(lo, hi):
        return [hi, hi - (hi - lo) / 3, lo + (hi - lo) / 3, lo]

    chart_data = {
        "climate_points_rainfall": _svg_points(climate_rainfall_norm, vmin=0, vmax=100),
        "climate_points_temp": _svg_points(climate_temp_norm, vmin=0, vmax=100),
        "climate_points_spi": _svg_points(drought_spi_norm, vmin=0, vmax=100),
        "climate_points_ndvi": _svg_points(ndvi_values_norm, vmin=0, vmax=100),
        "drought_points_spi": _svg_points(drought_spi, vmin=spi_min, vmax=spi_max),
        "drought_points_spei": _svg_points(drought_spei, vmin=spi_min, vmax=spi_max),
        "ndvi_points": _svg_points(ndvi_values, vmin=ndvi_min, vmax=ndvi_max),
        "chart_months": chart_months,
        "climate_labels": _labels(0, 100),
        "drought_labels": _labels(-3.0, 3.0),
        "ndvi_labels": _labels(0, 1.0),
    }

    demo_banner_level = _("High risk")
    demo_banner_body = _(
        "Current soil salinity levels in sector 4 are incompatible with durum wheat under drip irrigation. "
        "Consider switching to highly salt-tolerant crops or scheduling intensive leaching protocols before sowing."
    )
    latest_ra = RiskAssessment.objects.order_by("-created_at").first()
    if latest_ra:
        demo_banner_level = latest_ra.risk_level
        demo_banner_body = latest_ra.recommendation

    context = {
        "form": form,
        "result": result,
        "snapshot": snapshot,
        "region_map_config_json": json.dumps(_region_map_config()),
        "lakes_by_region_json": json.dumps(_lakes_by_region()),
        "dashboard_js_i18n_json": json.dumps(_dashboard_js_i18n()),
        "demo_risk_banner": {
            "level": demo_banner_level,
            "body": demo_banner_body,
        },
        **chart_data,
    }
    return render(request, "dashboard/dashboard.html", context)


# ============== METRICS MANAGEMENT VIEWS ==============

def data_ingestion(request):
    """Display data ingestion interface with drag-drop and manual entry options"""
    # Ensure reference data exists
    create_reference_data()
    
    context = {
        'regions': Region.objects.all(),
        'years': ObservationYear.objects.all(),
        'crops': CropType.objects.all(),
        'irrigation_methods': IrrigationMethod.objects.all(),
        'recent_imports': DataImportLog.objects.order_by('-import_date')[:5],
    }
    return render(request, 'dashboard/data_ingestion.html', context)


@require_http_methods(["POST"])
def api_submit_metrics(request):
    """API endpoint for manual data submission"""
    try:
        data = json.loads(request.body)
        metric_type = data.get('metric_type')
        
        # Get username
        username = request.user.username if request.user.is_authenticated else 'Anonymous'
        
        # Normalize common fields: accept names or ids
        # Region
        if data.get('region') and not data.get('region_id'):
            try:
                r = Region.objects.filter(name__iexact=data.get('region')).first()
                if r:
                    data['region_id'] = r.id
            except Exception:
                pass
        if data.get('region_id') and isinstance(data.get('region_id'), str) and data.get('region_id').isdigit():
            data['region_id'] = int(data['region_id'])

        # RBAC: viewers cannot submit data
        if _is_viewer_user(request.user):
            return JsonResponse({
                'success': False,
                'error': _('Viewers cannot submit data.'),
            }, status=403)

        # RBAC: subadmins may only submit data for their assigned region
        if _is_subadmin_user(request.user):
            allowed = request.user.profile.region_id
            submitted = data.get('region_id')
            if allowed and str(submitted) != str(allowed):
                return JsonResponse({
                    'success': False,
                    'error': _('You can only submit data for your assigned region.'),
                }, status=403)

        # Coordinate validation against region bounds
        region = Region.objects.filter(id=data.get('region_id')).first()
        if region:
            lat = data.get('latitude')
            lng = data.get('longitude')
            if lat is not None and lng is not None:
                is_valid, error_msg = _validate_coordinates_for_region(region, lat, lng)
                if not is_valid:
                    return JsonResponse({'success': False, 'error': error_msg}, status=400)

        # Year
        if data.get('year') and not data.get('year_id'):
            y = ObservationYear.objects.filter(label__iexact=data.get('year')).first()
            if y:
                data['year_id'] = y.id
        if data.get('year_id') and isinstance(data.get('year_id'), str) and data.get('year_id').isdigit():
            data['year_id'] = int(data['year_id'])

        # Crop
        if data.get('crop') and not data.get('crop_id'):
            c = CropType.objects.filter(name__iexact=data.get('crop')).first()
            if c:
                data['crop_id'] = c.id

        # Irrigation
        if data.get('irrigation') and not data.get('irrigation_id'):
            irr = IrrigationMethod.objects.filter(name__iexact=data.get('irrigation')).first()
            if irr:
                data['irrigation_id'] = irr.id

        # Timestamp alias
        if data.get('timestamp') and not data.get('measurement_date'):
            data['measurement_date'] = data.get('timestamp')

        # Create importer
        importer = DataImporter(source='Manual', username=username)

        # Submit the data
        result = importer.submit_form_data(metric_type, data)
        
        # Log the import
        import_log = importer.log_import(metric_type, notes=f'Manual entry via web form')
        action = getattr(importer, 'last_action', None)
        return JsonResponse({
            'success': True,
            'message': _('Data submitted successfully'),
            'records_imported': importer.imported_records,
            'log_id': import_log.id,
            'action': action,
        })
    
    except DataIngestionError as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
        }, status=400)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON format',
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Unexpected error: {str(e)}',
        }, status=500)


def parameter_review(request):
    """Display parameter review and manual entry interface with expandable categories"""
    context = {
        'regions': Region.objects.all(),
        'years': ObservationYear.objects.all(),
    }
    return render(request, 'dashboard/parameter_review.html', context)


def analysis_results(request):
    """Display analysis results and drought risk assessment"""
    import json
    from decimal import Decimal
    from django.core.serializers.json import DjangoJSONEncoder

    def safe_float(val):
        if val is None: return None
        try: return float(val)
        except: return None

    # Default to Bizerte for initial data
    default_region = Region.objects.filter(name="Bizerte").first() or Region.objects.order_by('name').first()

    # Latest metrics from each table, filtered to Bizerte
    latest_soil = SoilMetrics.objects.filter(region=default_region).order_by('-measurement_date').first()
    latest_climate = ClimateMetrics.objects.filter(region=default_region).order_by('-measurement_date').first()
    latest_remote = RemoteSensingMetrics.objects.filter(region=default_region).order_by('-measurement_date').first()
    latest_hydro = HydrologyMetrics.objects.filter(region=default_region).order_by('-measurement_date').first()
    latest_agri = AgriculturalMetrics.objects.filter(region=default_region).order_by('-measurement_date').first()

    latest_prediction = DroughtPrediction.objects.filter(region=default_region).order_by('-generated_at').select_related('region', 'year').first()

    # Extended initial data for metrics
    initial_data = {
        'soil_moisture_pct': safe_float(getattr(latest_soil, 'moisture_content_percent', None)),
        'rainfall_mm': safe_float(getattr(latest_climate, 'rainfall_mm', None)),
        'ndvi': safe_float(getattr(latest_remote, 'ndvi', None)),
        'etc_mm': safe_float(getattr(latest_climate, 'evapotranspiration_etc_mmday', None)),
        'groundwater_m': safe_float(getattr(latest_hydro, 'groundwater_depth_m', None)),
        'stress_pct': safe_float(getattr(latest_agri, 'yield_reduction_factor', None)),
        'temperature_c': safe_float(getattr(latest_climate, 'temperature_mean_c', None)),
        'wind_speed': safe_float(getattr(latest_climate, 'wind_speed_ms', None)),
        'humidity': safe_float(getattr(latest_climate, 'relative_humidity_percent', None)),
        'solar_radiation': safe_float(getattr(latest_climate, 'solar_radiation_mjm2day', None)),
    }

    # ── Region / Year maps for JS ──
    region_map = {r.name.lower(): r.id for r in Region.objects.all()}
    year_map = {y.label.split()[0]: y.id for y in ObservationYear.objects.all() if y.label.split()[0].isdigit()}
    default_region_id = region_map.get(default_region.name.lower()) if default_region else 1
    default_year_id = year_map.get(str(timezone.now().year), 1)

    # ── Soil water balance bars (last 6 records for Bizerte) ──
    soil_qs = list(SoilMetrics.objects.filter(region=default_region).order_by('-measurement_date')[:6])
    soil_qs.reverse()
    soil_water_vals = [safe_float(s.moisture_content_percent) or 0 for s in soil_qs]
    # Normalize to field capacity (100% = field capacity, not saturation)
    fc_soil = SoilMetrics.objects.filter(region=default_region, field_capacity_percent__isnull=False).order_by('-measurement_date').first()
    fc_val = float(fc_soil.field_capacity_percent) if fc_soil and fc_soil.field_capacity_percent else 30.0
    soil_water_bars = [min(1.0, v / fc_val) if v > 0 else 0 for v in soil_water_vals]
    while len(soil_water_bars) < 6:
        soil_water_bars.append(0)

    # ── Drought risk trend SVG path ──
    risk_today_raw = safe_float(getattr(latest_prediction, 'current_risk_score', None))
    risk_7_raw = safe_float(getattr(latest_prediction, 'risk_7day', None))
    risk_30_raw = safe_float(getattr(latest_prediction, 'risk_30day', None))
    risk_today = risk_today_raw if risk_today_raw is not None else 50
    risk_7 = risk_7_raw if risk_7_raw is not None else 60
    risk_30 = risk_30_raw if risk_30_raw is not None else 70
    risk_pts = [
        max(0, risk_today - 20), max(0, risk_today - 12), max(0, risk_today - 6),
        risk_today, min(100, risk_7), min(100, (risk_7 + risk_30) / 2), min(100, risk_30),
    ]

    def _risk_svg(pts, svg_w=400, svg_h=200, pad_l=29, pad_r=29, pad_t=0, pad_b=0):
        cw = svg_w - pad_l - pad_r
        ch = svg_h - pad_t - pad_b
        n = len(pts)
        if n == 0:
            return ""
        coords = []
        for i, v in enumerate(pts):
            x = pad_l + (i / (n - 1)) * cw
            y = pad_t + ch - (v / 100) * ch
            coords.append((x, y))
        d = f"M{coords[0][0]},{coords[0][1]}"
        for i in range(1, n):
            px, py = coords[i - 1]
            cx, cy = (px + coords[i][0]) / 2, py
            d += f" Q{cx},{cy} {coords[i][0]},{coords[i][1]}"
        fill = d + f" L{coords[-1][0]},{pad_b + ch} L{coords[0][0]},{pad_b + ch} Z"
        return d, fill

    trend_line, trend_fill = _risk_svg(risk_pts)

    def _time_ago(date_val):
        if date_val is None:
            return "Unknown"
        from datetime import date as dt_date
        if hasattr(date_val, 'date'):
            date_val = date_val.date()
        delta = (timezone.now().date() - date_val).days
        if delta == 0:
            return "Today"
        if delta == 1:
            return "Yesterday"
        return f"{delta}d ago"

    # ── Dynamic alerts from data ──
    alerts = []
    sm = safe_float(getattr(latest_soil, 'moisture_content_percent', None))
    if sm is not None and sm < 12:
        sm_date = getattr(latest_soil, 'measurement_date', None)
        alerts.append({
            'priority': 'High Priority', 'priority_cls': 'border-error bg-error-container/20 text-error',
            'time': _time_ago(sm_date), 'title': 'Soil Moisture Critical Drop',
            'desc': f'Sensor station S-04 reported &lt; {sm:.0f}% moisture. Immediate verification required.',
            'btn_text': 'Action Required', 'btn_cls': 'text-error',
        })
    rf_30d = safe_float(getattr(latest_climate, 'rainfall_mm', None))
    if rf_30d is not None and rf_30d < 20:
        rf_date = getattr(latest_climate, 'measurement_date', None)
        alerts.append({
            'priority': 'Medium Priority', 'priority_cls': 'border-secondary bg-surface-container-low text-secondary',
            'time': _time_ago(rf_date), 'title': 'Precipitation Gap Warning',
            'desc': f'Low rainfall ({rf_30d:.1f}mm recorded). Trend analysis updated.',
            'btn_text': 'View Trends', 'btn_cls': 'text-secondary',
        })
    ndvi_val = safe_float(getattr(latest_remote, 'ndvi', None))
    if ndvi_val is not None and ndvi_val < 0.3:
        ndvi_date = getattr(latest_remote, 'measurement_date', None)
        alerts.append({
            'priority': 'Medium Priority', 'priority_cls': 'border-yellow-500 bg-yellow-50 text-yellow-700',
            'time': _time_ago(ndvi_date), 'title': 'Vegetation Stress Detected',
            'desc': f'NDVI at {ndvi_val:.2f} indicates significant vegetation water stress in the basin.',
            'btn_text': 'Analyze', 'btn_cls': 'text-yellow-600',
        })
    if not alerts:
        alerts.append({
            'priority': 'Info', 'priority_cls': 'border-primary bg-primary-container/10 text-primary',
            'time': 'Just now', 'title': 'All Clear',
            'desc': 'No critical thresholds exceeded. Standard monitoring continues.',
            'btn_text': 'Dashboard', 'btn_cls': 'text-primary',
        })

    # ── Soil water bar labels (from measurement dates) ──
    soil_labels = []
    for s in soil_qs:
        d = getattr(s, 'measurement_date', None)
        if d:
            soil_labels.append(d.strftime('%b %d'))
        else:
            soil_labels.append('—')
    while len(soil_labels) < 6:
        soil_labels.append('—')

    # ── Trend chart labels ──
    trend_labels = ['T-21', 'T-14', 'T-7', 'Today', '+7 Days', '+14 Days', '+30 Days']

    # ── Heatmap zone areas (proportional to current risk) ──
    base_area = 1200
    extreme_pct = max(0.05, (risk_today / 100) * 0.35)
    severe_pct = max(0.05, (risk_7 / 100) * 0.30)
    moderate_pct = max(0.05, 0.25 - (risk_today / 100) * 0.10)
    safe_pct = max(0.05, 1.0 - extreme_pct - severe_pct - moderate_pct)
    total_pct = extreme_pct + severe_pct + moderate_pct + safe_pct
    # Normalize
    extreme_pct /= total_pct; severe_pct /= total_pct; moderate_pct /= total_pct; safe_pct /= total_pct

    heatmap_zones = {
        'extreme_km2': f"{extreme_pct * base_area:.1f}",
        'severe_km2': f"{severe_pct * base_area:.1f}",
        'moderate_km2': f"{moderate_pct * base_area:.1f}",
        'total_km2': f"{base_area:,}",
    }

    # Year choices for template
    year_choices = [(y.label.split()[0], y.label) for y in ObservationYear.objects.all() if y.label.split()[0].isdigit()]

    context = {
        'regions': Region.objects.all(),
        'years': year_choices,
        'now_year': str(timezone.now().year),
        'initial_data_json': json.dumps(initial_data, cls=DjangoJSONEncoder),
        'region_map_json': json.dumps(region_map),
        'year_map_json': json.dumps(year_map),
        'default_region_id': default_region_id,
        'default_region_name': default_region.name if default_region else 'Bizerte',
        'default_year_id': default_year_id,
        'soil_water_bars': json.dumps(soil_water_bars),
        'risk_pts_json': json.dumps(risk_pts),
        'trend_line': trend_line,
        'trend_fill': trend_fill,
        'trend_labels': trend_labels,
        'alerts': alerts,
        'heatmap_zones': heatmap_zones,
        'soil_water_labels': soil_labels,
    }
    return render(request, 'dashboard/analysis.html', context)


def analysis_export_csv(request):
    """Export a comprehensive analysis report for a region/year as PDF."""
    from datetime import date
    from decimal import Decimal
    from django.http import HttpResponse
    from fpdf import FPDF

    region_id = request.GET.get('region_id')
    year_id = request.GET.get('year_id')

    if not region_id or not year_id:
        return HttpResponse('Missing region_id or year_id', status=400)

    try:
        region = Region.objects.get(pk=region_id)
        year = ObservationYear.objects.get(pk=year_id)
    except (Region.DoesNotExist, ObservationYear.DoesNotExist):
        return HttpResponse('Invalid region or year', status=404)

    from .models import (
        DroughtPrediction, SoilMetrics, ClimateMetrics,
        RemoteSensingMetrics, HydrologyMetrics, AgriculturalMetrics,
        DroughtIndices
    )

    try:
        from .prediction_engine.pipeline import DroughtPredictionPipeline
        pipeline = DroughtPredictionPipeline()
        result = pipeline.predict_for_region(region, year, use_llm=False)
        pred = pipeline.save_prediction(region, year, result)
    except Exception:
        pred = DroughtPrediction.objects.filter(region=region, year=year).order_by('-generated_at').first()

    latest_soil = SoilMetrics.objects.filter(region=region, year=year).order_by('-measurement_date').first()
    latest_climate = ClimateMetrics.objects.filter(region=region, year=year).order_by('-measurement_date').first()
    latest_rs = RemoteSensingMetrics.objects.filter(region=region, year=year).order_by('-measurement_date').first()
    latest_hydro = HydrologyMetrics.objects.filter(region=region, year=year).order_by('-measurement_date').first()
    latest_agri = AgriculturalMetrics.objects.filter(region=region, year=year).order_by('-measurement_date').first()
    latest_di = DroughtIndices.objects.filter(region=region, year=year).order_by('-measurement_date').first()

    today_str = date.today().isoformat()
    risk_today = float(str(pred.current_risk_score)) if pred and pred.current_risk_score is not None else None
    risk_7 = float(str(pred.risk_7day)) if pred and pred.risk_7day is not None else None
    risk_30 = float(str(pred.risk_30day)) if pred and pred.risk_30day is not None else None

    def _risk_label(s):
        if s is None:
            return 'N/A'
        if s < 25:
            return 'Safe'
        if s < 50:
            return 'Watch'
        if s < 75:
            return 'Severe'
        return 'Extreme'

    drivers = pred.drivers if pred and pred.drivers else {}
    explanation = pred.explanation if pred and pred.explanation else ''

    def _fmt(val):
        if val is None:
            return '-'
        if isinstance(val, (Decimal, float, int)):
            return str(round(float(val), 2))
        return str(val)

    def _pdf_str(s):
        if not s:
            return ''
        try:
            s.encode('latin-1')
            return s
        except UnicodeEncodeError:
            import unicodedata
            return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')

    # ── Build PDF ──
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    w = pdf.w - pdf.l_margin - pdf.r_margin

    # Title
    pdf.set_font('Helvetica', 'B', 18)
    pdf.cell(w, 12, 'Drought Analysis Report', border=0, align='C')
    pdf.ln(14)

    # ── 1. Report Summary ──
    pdf.set_font('Helvetica', 'B', 13)
    pdf.set_fill_color(230, 240, 255)
    pdf.cell(w, 8, '1. Report Summary', border=0, fill=True)
    pdf.ln(10)

    pdf.set_font('Helvetica', '', 10)
    summary_rows = [
        ('Region', region.name),
        ('Year', year.label),
        ('Export Date', today_str),
        ('Overall Risk', f'{_risk_label(risk_today)} ({_fmt(risk_today)}/100)'),
        ('7-Day Forecast', f'{_risk_label(risk_7)} ({_fmt(risk_7)}/100)'),
        ('30-Day Forecast', f'{_risk_label(risk_30)} ({_fmt(risk_30)}/100)'),
    ]
    for label, value in summary_rows:
        pdf.set_font('Helvetica', 'B', 10)
        pdf.cell(50, 7, label, border=0)
        pdf.set_font('Helvetica', '', 10)
        pdf.cell(w - 50, 7, _pdf_str(value), border=0)
        pdf.ln(7)

    pdf.ln(4)

    # ── 2. Key Drought Drivers ──
    pdf.set_font('Helvetica', 'B', 13)
    pdf.set_fill_color(230, 240, 255)
    pdf.cell(w, 8, '2. Key Drought Drivers', border=0, fill=True)
    pdf.ln(10)

    # Table header
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_fill_color(200, 215, 240)
    col_w = [w * 0.45, w * 0.25, w * 0.30]
    headers = ['Driver', 'Contribution (%)', 'Score (0-100)']
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h, border=1, fill=True, align='C')
    pdf.ln()

    driver_keys = [
        ('rainfall_deficit', 'Rainfall Deficit'),
        ('high_temperature', 'High Temperature'),
        ('soil_moisture_decline', 'Soil Moisture Decline'),
        ('vegetation_stress', 'Vegetation Stress'),
        ('high_evapotranspiration', 'High Evapotranspiration'),
    ]
    abs_drivers = drivers.get('absolute', {}) if isinstance(drivers, dict) else {}
    pdf.set_font('Helvetica', '', 9)
    for i, (key, label) in enumerate(driver_keys):
        contrib = drivers.get(key, '') if isinstance(drivers, dict) else ''
        abs_val = abs_drivers.get(key, '') if isinstance(abs_drivers, dict) else ''
        fill = (i % 2 == 0)
        if fill:
            pdf.set_fill_color(245, 248, 252)
        pdf.cell(col_w[0], 6, f'  {label}', border=1, fill=fill)
        pdf.cell(col_w[1], 6, _fmt(contrib), border=1, fill=fill, align='C')
        pdf.cell(col_w[2], 6, _fmt(abs_val), border=1, fill=fill, align='C')
        pdf.ln()

    pdf.ln(6)

    # ── 3. Drought Insight ──
    pdf.set_font('Helvetica', 'B', 13)
    pdf.set_fill_color(230, 240, 255)
    pdf.cell(w, 8, '3. Drought Insight', border=0, fill=True)
    pdf.ln(10)

    pdf.set_font('Helvetica', '', 10)
    pdf.set_fill_color(252, 252, 252)
    pdf.multi_cell(w, 6, _pdf_str(explanation) if explanation else 'No insight available.', border=0, fill=True)
    pdf.ln(6)

    # ── 4. Current Metrics Snapshot ──
    pdf.set_font('Helvetica', 'B', 13)
    pdf.set_fill_color(230, 240, 255)
    pdf.cell(w, 8, '4. Current Metrics Snapshot', border=0, fill=True)
    pdf.ln(10)

    snapshots = [
        ('Soil', 'Moisture (%)', latest_soil.moisture_content_percent if latest_soil else None),
        ('Soil', 'Field Capacity (%)', latest_soil.field_capacity_percent if latest_soil else None),
        ('Soil', 'Wilting Point (%)', latest_soil.wilting_point_percent if latest_soil else None),
        ('Soil', 'Sand Ratio', latest_soil.sand_ratio if latest_soil else None),
        ('Soil', 'Clay Ratio', latest_soil.clay_ratio if latest_soil else None),
        ('Soil', 'Organic Matter (%)', latest_soil.organic_matter_percent if latest_soil else None),
        ('Climate', 'Rainfall (mm)', latest_climate.rainfall_mm if latest_climate else None),
        ('Climate', 'Max Temperature (C)', latest_climate.temperature_max_c if latest_climate else None),
        ('Climate', 'Min Temperature (C)', latest_climate.temperature_min_c if latest_climate else None),
        ('Climate', 'Mean Temperature (C)', latest_climate.temperature_mean_c if latest_climate else None),
        ('Climate', 'Humidity (%)', latest_climate.relative_humidity_percent if latest_climate else None),
        ('Climate', 'Wind Speed (m/s)', latest_climate.wind_speed_ms if latest_climate else None),
        ('Climate', 'ET0 (mm/day)', latest_climate.evapotranspiration_et0_mmday if latest_climate else None),
        ('Climate', 'ETc (mm/day)', latest_climate.evapotranspiration_etc_mmday if latest_climate else None),
        ('Drought Indices', 'SPI (1-month)', latest_di.spi_1month if latest_di else None),
        ('Drought Indices', 'SPEI (1-month)', latest_di.spei_1month if latest_di else None),
        ('Drought Indices', 'PDSI', latest_di.pdsi_value if latest_di else None),
        ('Remote Sensing', 'NDVI', latest_rs.ndvi if latest_rs else None),
        ('Remote Sensing', 'LST (C)', latest_rs.land_surface_temperature_c if latest_rs else None),
        ('Remote Sensing', 'Satellite SM (%)', latest_rs.satellite_soil_moisture_percent if latest_rs else None),
        ('Hydrology', 'Precipitation (mm)', latest_hydro.precipitation_mm if latest_hydro else None),
        ('Hydrology', 'Groundwater Depth (m)', latest_hydro.groundwater_depth_m if latest_hydro else None),
        ('Hydrology', 'Runoff (mm)', latest_hydro.runoff_mm if latest_hydro else None),
        ('Hydrology', 'River Flow (m3/s)', latest_hydro.river_flow_m3s if latest_hydro else None),
        ('Agriculture', 'Growth Stage', latest_agri.growth_stage.name if latest_agri and hasattr(latest_agri.growth_stage, 'name') else (latest_agri.growth_stage if latest_agri else None)),
        ('Agriculture', 'Crop Coefficient (Kc)', latest_agri.crop_coefficient_kc if latest_agri else None),
        ('Agriculture', 'Water Requirement (mm/day)', latest_agri.crop_water_requirement_mmday if latest_agri else None),
        ('Agriculture', 'Yield Reduction Factor', latest_agri.yield_reduction_factor if latest_agri else None),
    ]

    # Snapshot table — 3 columns: Category, Metric, Value
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_fill_color(200, 215, 240)
    sc_w = [w * 0.22, w * 0.52, w * 0.26]
    for i, h in enumerate(['Category', 'Metric', 'Value']):
        pdf.cell(sc_w[i], 7, h, border=1, fill=True, align='C')
    pdf.ln()

    pdf.set_font('Helvetica', '', 8)
    for i, (cat, metric, val) in enumerate(snapshots):
        fill = (i % 2 == 0)
        if fill:
            pdf.set_fill_color(245, 248, 252)
        pdf.cell(sc_w[0], 5, f'  {cat}', border=1, fill=fill)
        pdf.cell(sc_w[1], 5, f'  {metric}', border=1, fill=fill)
        pdf.cell(sc_w[2], 5, _fmt(val), border=1, fill=fill, align='C')
        pdf.ln()

    pdf.ln(6)

    # ── 5. Daily Time Series note ──
    pdf.set_font('Helvetica', 'B', 13)
    pdf.set_fill_color(230, 240, 255)
    pdf.cell(w, 8, '5. Daily Time Series Data', border=0, fill=True)
    pdf.ln(10)

    pdf.set_font('Helvetica', '', 10)
    pdf.multi_cell(w, 6, 'The full daily time series with all metric columns and risk scores is included in the attached CSV file (same filename with .csv extension).', border=0)
    pdf.ln(4)

    # Footer
    pdf.set_font('Helvetica', 'I', 8)
    pdf.cell(w, 5, f'Generated by ABC Basin Drought Early Warning System on {today_str}', align='C')

    # ── Build response ──
    from io import BytesIO
    buffer = BytesIO()
    pdf.output(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    fname = f'analysis_{region.name.lower()}_{year.label.split()[0]}'
    response['Content-Disposition'] = f'attachment; filename="{fname}.pdf"'
    return response


# ── Registry of all available chartable metrics ──
METRICS_REGISTRY = [
    # (key, model_class, field, label, unit, category)
    # Soil
    ('soil_moisture',     'SoilMetrics', 'moisture_content_percent',      'Soil Moisture',          '%',     'Soil'),
    ('field_capacity',    'SoilMetrics', 'field_capacity_percent',        'Field Capacity',         '%',     'Soil'),
    ('wilting_point',     'SoilMetrics', 'wilting_point_percent',         'Wilting Point',          '%',     'Soil'),
    ('sand_ratio',        'SoilMetrics', 'sand_ratio',                    'Sand Ratio',             '%',     'Soil'),
    ('clay_ratio',        'SoilMetrics', 'clay_ratio',                    'Clay Ratio',             '%',     'Soil'),
    ('organic_matter',    'SoilMetrics', 'organic_matter_percent',        'Organic Matter',         '%',     'Soil'),
    ('awc',               None,          None,                            'Available Water Capacity','mm',   'Soil'),  # derived
    # Climate
    ('rainfall',          'ClimateMetrics', 'rainfall_mm',                'Rainfall',               'mm',    'Climate'),
    ('temp_max',          'ClimateMetrics', 'temperature_max_c',          'Max Temperature',        '°C',    'Climate'),
    ('temp_min',          'ClimateMetrics', 'temperature_min_c',          'Min Temperature',        '°C',    'Climate'),
    ('temp_mean',         'ClimateMetrics', 'temperature_mean_c',         'Mean Temperature',       '°C',    'Climate'),
    ('humidity',          'ClimateMetrics', 'relative_humidity_percent',  'Humidity',               '%',     'Climate'),
    ('wind_speed',        'ClimateMetrics', 'wind_speed_ms',              'Wind Speed',             'm/s',   'Climate'),
    ('solar_radiation',   'ClimateMetrics', 'solar_radiation_mjm2day',    'Solar Radiation',        'MJ/m²','Climate'),
    ('et0',               'ClimateMetrics', 'evapotranspiration_et0_mmday','ET₀ (Reference)',       'mm/day','Climate'),
    ('etc',               'ClimateMetrics', 'evapotranspiration_etc_mmday','ETc (Crop)',            'mm/day','Climate'),
    # Drought Indices
    ('spi1',              'DroughtIndices', 'spi_1month',                 'SPI (1-month)',          '',      'Drought'),
    ('spi3',              'DroughtIndices', 'spi_3month',                 'SPI (3-month)',          '',      'Drought'),
    ('spi12',             'DroughtIndices', 'spi_12month',                'SPI (12-month)',         '',      'Drought'),
    ('spei1',             'DroughtIndices', 'spei_1month',                'SPEI (1-month)',         '',      'Drought'),
    ('spei3',             'DroughtIndices', 'spei_3month',                'SPEI (3-month)',         '',      'Drought'),
    ('spei12',            'DroughtIndices', 'spei_12month',               'SPEI (12-month)',        '',      'Drought'),
    ('pdsi',              'DroughtIndices', 'pdsi_value',                 'PDSI',                   '',      'Drought'),
    # Remote Sensing
    ('ndvi',              'RemoteSensingMetrics', 'ndvi',                 'NDVI',                   '',      'Remote Sensing'),
    ('ndwi',              'RemoteSensingMetrics', 'ndwi',                 'NDWI',                   '',      'Remote Sensing'),
    ('lst',               'RemoteSensingMetrics', 'land_surface_temperature_c', 'LST',              '°C',    'Remote Sensing'),
    ('sat_soil_moisture', 'RemoteSensingMetrics', 'satellite_soil_moisture_percent', 'Satellite SM','%',    'Remote Sensing'),
    ('vci',               'RemoteSensingMetrics', 'vegetation_condition_index', 'VCI',              '%',     'Remote Sensing'),
    # Hydrology
    ('precip',            'HydrologyMetrics', 'precipitation_mm',         'Precipitation',          'mm',    'Hydrology'),
    ('et_hydro',          'HydrologyMetrics', 'evapotranspiration_mm',    'Evapotranspiration',     'mm',    'Hydrology'),
    ('groundwater',       'HydrologyMetrics', 'groundwater_depth_m',      'Groundwater Depth',      'm',     'Hydrology'),
    ('runoff',            'HydrologyMetrics', 'runoff_mm',                'Runoff',                 'mm',    'Hydrology'),
    ('river_flow',        'HydrologyMetrics', 'river_flow_m3s',           'River Flow',             'm³/s',  'Hydrology'),
    ('reservoir',         'HydrologyMetrics', 'reservoir_storage_m3',     'Reservoir Storage',      'm³',    'Hydrology'),
    ('water_balance',     'HydrologyMetrics', 'water_balance_percent',    'Water Balance',          '%',     'Hydrology'),
    # Agriculture
    ('kc',                'AgriculturalMetrics', 'crop_coefficient_kc',   'Crop Coefficient (Kc)',  '',      'Agriculture'),
    ('cwr',               'AgriculturalMetrics', 'crop_water_requirement_mmday', 'Water Requirement', 'mm/day','Agriculture'),
    ('yield_reduction',   'AgriculturalMetrics', 'yield_reduction_factor','Yield Reduction',        '0-1',   'Agriculture'),
    ('irrigation_eff',    'AgriculturalMetrics', 'irrigation_efficiency_percent', 'Irrigation Eff.','%',    'Agriculture'),
    ('leaf_temp',         'AgriculturalMetrics', 'leaf_temperature_c',    'Leaf Temperature',       '°C',    'Agriculture'),
]

MODEL_MAP = {
    'SoilMetrics': SoilMetrics,
    'ClimateMetrics': ClimateMetrics,
    'DroughtIndices': DroughtIndices,
    'RemoteSensingMetrics': RemoteSensingMetrics,
    'HydrologyMetrics': HydrologyMetrics,
    'AgriculturalMetrics': AgriculturalMetrics,
}


def calculations_view(request):
    """Display calculated metrics, formulas, and custom chart builder."""
    rid = request.GET.get('region_id')
    region = get_object_or_404(Region, pk=rid) if rid else Region.objects.filter(name='Bizerte').first() or Region.objects.first()
    snap_date = request.GET.get('snap_date')

    from decimal import Decimal
    from datetime import date

    def _latest_with_date(model_cls, region, snap_date):
        filters = {'region': region}
        if snap_date:
            filters['measurement_date__lte'] = snap_date
        return model_cls.objects.filter(**filters).order_by('-measurement_date').first()

    latest_soil = _latest_with_date(SoilMetrics, region, snap_date)
    latest_climate = _latest_with_date(ClimateMetrics, region, snap_date)
    latest_di = _latest_with_date(DroughtIndices, region, snap_date)
    latest_rs = _latest_with_date(RemoteSensingMetrics, region, snap_date)
    latest_hydro = _latest_with_date(HydrologyMetrics, region, snap_date)
    latest_agri = _latest_with_date(AgriculturalMetrics, region, snap_date)

    def _f(v):
        if v is None:
            return '—'
        if isinstance(v, Decimal):
            return round(float(v), 2)
        return v

    mean_temp = None
    if latest_climate and latest_climate.temperature_max_c is not None and latest_climate.temperature_min_c is not None:
        try:
            mean_temp = round((float(latest_climate.temperature_max_c) + float(latest_climate.temperature_min_c)) / 2.0, 2)
        except Exception:
            mean_temp = None

    awc = None
    awc_formula_parts = []
    if latest_soil:
        fc = latest_soil.field_capacity_percent
        wp = latest_soil.wilting_point_percent
        rz = latest_soil.root_zone_depth_mm
        if fc is not None and wp is not None and rz:
            try:
                fc_f = float(fc)
                wp_f = float(wp)
                awc = round(((fc_f - wp_f) / 100.0) * float(rz), 2)
                awc_formula_parts = [f'FC={fc_f}%', f'WP={wp_f}%', f'RZ={rz}mm', f'({fc_f}-{wp_f})/100×{rz}={awc}mm']
            except Exception:
                pass

    soil_water_deficit = None
    if latest_soil and latest_climate:
        sm = latest_soil.moisture_content_percent
        fc_v = latest_soil.field_capacity_percent
        if sm is not None and fc_v is not None and float(fc_v) > 0:
            soil_water_deficit = round(max(0, 100 - (float(sm) / float(fc_v) * 100)), 2)

    et_def_calc = None
    if latest_climate and latest_agri:
        et0 = latest_climate.evapotranspiration_et0_mmday
        kc = latest_agri.crop_coefficient_kc
        if et0 is not None and kc is not None:
            et_def_calc = f'{_f(et0)} × {_f(kc)} = {round(float(et0)*float(kc), 2)} mm/day'

    calc_sections = [
        {
            'category': 'Soil',
            'items': [
                {'name': 'Field Capacity (%)',  'formula': 'Saxton & Rawls PTF: 0.28 + 0.57×clay − 0.24×sand', 'value': _f(latest_soil.field_capacity_percent if latest_soil else None)},
                {'name': 'Wilting Point (%)',   'formula': 'Saxton & Rawls PTF: 0.06 + 0.42×clay − 0.07×sand', 'value': _f(latest_soil.wilting_point_percent if latest_soil else None)},
                {'name': 'AWC (mm)',             'formula': awc_formula_parts[3] if awc_formula_parts else '(FC − WP)/100 × Root Zone Depth', 'value': _f(awc)},
                {'name': 'Soil Water Deficit (%)','formula': 'max(0, 100 − (SM/FC × 100))', 'value': _f(soil_water_deficit)},
                {'name': 'Organic Matter (%)',   'formula': 'Loss on ignition (LOI) method', 'value': _f(latest_soil.organic_matter_percent if latest_soil else None)},
                {'name': 'Infiltration (mm/hr)', 'formula': 'Double-ring infiltrometer or PTF from texture', 'value': _f(latest_soil.infiltration_rate_mmhr if latest_soil else None)},
            ]
        },
        {
            'category': 'Climate',
            'items': [
                {'name': 'Mean Temp (°C)',       'formula': f'({_f(latest_climate.temperature_max_c if latest_climate else None)} + {_f(latest_climate.temperature_min_c if latest_climate else None)}) ÷ 2', 'value': _f(mean_temp)},
                {'name': 'ET₀ (mm/day)',          'formula': 'Penman-Monteith FAO-56: (0.408ΔRn + γ·(900/(T+273))·U₂·(es−ea)) / (Δ+γ(1+0.34U₂))', 'value': _f(latest_climate.evapotranspiration_et0_mmday if latest_climate else None)},
                {'name': 'ETc (mm/day)',          'formula': f'ET₀ × Kc: {et_def_calc or "ET₀ × Kc (FAO-56 lookup)"}', 'value': _f(latest_climate.evapotranspiration_etc_mmday if latest_climate else None)},
                {'name': 'Rainfall (mm)',         'formula': 'Station / satellite measurement (CHIRPS, gauge)', 'value': _f(latest_climate.rainfall_mm if latest_climate else None)},
                {'name': 'Humidity (%)',          'formula': 'Station measurement (DHT22 / hygrometer)', 'value': _f(latest_climate.relative_humidity_percent if latest_climate else None)},
                {'name': 'Wind Speed (m/s)',      'formula': 'Station measurement (anemometer)', 'value': _f(latest_climate.wind_speed_ms if latest_climate else None)},
            ]
        },
        {
            'category': 'Drought Indices',
            'items': [
                {'name': 'SPI (1-month)',         'formula': 'Gamma(α,β) fitted to 30d rainfall → inverse normal CDF', 'value': _f(latest_di.spi_1month if latest_di else None)},
                {'name': 'SPI (3-month)',         'formula': 'Gamma(α,β) fitted to 90d rainfall → inverse normal CDF', 'value': _f(latest_di.spi_3month if latest_di else None)},
                {'name': 'SPI (12-month)',        'formula': 'Gamma(α,β) fitted to 365d rainfall → inverse normal CDF', 'value': _f(latest_di.spi_12month if latest_di else None)},
                {'name': 'SPEI (1-month)',        'formula': 'Water balance (P − ET₀) anomaly ÷ historical σ', 'value': _f(latest_di.spei_1month if latest_di else None)},
                {'name': 'PDSI',                  'formula': 'Palmer model: soil water balance with climatic coefficient', 'value': _f(latest_di.pdsi_value if latest_di else None)},
            ]
        },
        {
            'category': 'Remote Sensing',
            'items': [
                {'name': 'NDVI',                  'formula': '(NIR − Red) / (NIR + Red)', 'value': _f(latest_rs.ndvi if latest_rs else None)},
                {'name': 'LST (°C)',              'formula': 'Split-window algorithm from TIR bands', 'value': _f(latest_rs.land_surface_temperature_c if latest_rs else None)},
                {'name': 'Satellite SM (%)',      'formula': 'SMAP L-band radiometer (0-5 cm depth)', 'value': _f(latest_rs.satellite_soil_moisture_percent if latest_rs else None)},
                {'name': 'VCI (%)',               'formula': '(NDVI − NDVI_min) / (NDVI_max − NDVI_min) × 100', 'value': _f(latest_rs.vegetation_condition_index if latest_rs else None)},
            ]
        },
        {
            'category': 'Hydrology',
            'items': [
                {'name': 'Precipitation (mm)',    'formula': 'Station / CHIRPS measurement', 'value': _f(latest_hydro.precipitation_mm if latest_hydro else None)},
                {'name': 'Groundwater Depth (m)', 'formula': 'Well measurement / piezometer', 'value': _f(latest_hydro.groundwater_depth_m if latest_hydro else None)},
                {'name': 'Runoff (mm)',            'formula': 'SCS-CN: Q = (P−0.2S)²/(P+0.8S) with S=(25400/CN)−254', 'value': _f(latest_hydro.runoff_mm if latest_hydro else None)},
                {'name': 'River Flow (m³/s)',     'formula': 'Stage-discharge rating curve (Q = a·h^b)', 'value': _f(latest_hydro.river_flow_m3s if latest_hydro else None)},
                {'name': 'Water Balance (%)',     'formula': '(Supply − Demand) / Demand × 100', 'value': _f(latest_hydro.water_balance_percent if latest_hydro else None)},
            ]
        },
        {
            'category': 'Agriculture',
            'items': [
                {'name': 'Crop Coeff. (Kc)',      'formula': 'FAO-56 growth-stage lookup table', 'value': _f(latest_agri.crop_coefficient_kc if latest_agri else None)},
                {'name': 'CWR (mm/day)',           'formula': 'ETc = ET₀ × Kc', 'value': _f(latest_agri.crop_water_requirement_mmday if latest_agri else None)},
                {'name': 'Yield Reduction (0-1)',  'formula': '1 − (actual / potential) yield', 'value': _f(latest_agri.yield_reduction_factor if latest_agri else None)},
                {'name': 'Irrigation Efficiency (%)','formula': 'Applied water beneficially used ÷ total applied', 'value': _f(latest_agri.irrigation_efficiency_percent if latest_agri else None)},
            ]
        },
    ]

    import json as _json
    context = {
        'regions': Region.objects.all(),
        'years': ObservationYear.objects.all(),
        'calc_sections': calc_sections,
        'metrics_json': _json.dumps([{'key': k, 'label': f'{cat} – {lbl}', 'model': m, 'field': f, 'unit': u, 'category': cat} for k, m, f, lbl, u, cat in METRICS_REGISTRY if m is not None]),
        'selected_region': region,
        'snap_date': snap_date or '',
    }
    return render(request, 'dashboard/calculations.html', context)


def calculations_chart_api(request):
    """Return time-series data for a chosen metric + region + date range."""
    from datetime import datetime, date
    from decimal import Decimal
    import json

    region_id = request.GET.get('region_id')
    metric = request.GET.get('metric')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not all([region_id, metric]):
        return JsonResponse({'error': 'region_id and metric required'}, status=400)

    region = get_object_or_404(Region, pk=region_id)

    # Look up metric in registry
    reg_entry = None
    for entry in METRICS_REGISTRY:
        if entry[0] == metric:
            reg_entry = entry
            break
    if not reg_entry:
        return JsonResponse({'error': f'Unknown metric: {metric}'}, status=400)

    _key, model_name, field_name, label, unit, cat = reg_entry

    if model_name is None:
        # derived metric — not available as raw time-series
        return JsonResponse({'error': 'Derived metrics not available as time-series'}, status=400)

    model_cls = MODEL_MAP[model_name]

    filters = {'region': region}
    if date_from:
        filters['measurement_date__gte'] = date_from
    if date_to:
        filters['measurement_date__lte'] = date_to

    qs = model_cls.objects.filter(**filters).order_by('measurement_date').values('measurement_date', field_name)

    dates = []
    values = []
    for row in qs:
        v = row[field_name]
        if v is not None:
            if isinstance(v, Decimal):
                v = float(v)
            dates.append(row['measurement_date'].isoformat() if hasattr(row['measurement_date'], 'isoformat') else str(row['measurement_date']))
            values.append(v)

    return JsonResponse({
        'metric': metric,
        'label': label,
        'unit': unit,
        'category': cat,
        'dates': dates,
        'values': values,
    })


def calculations_data(request):
    """Return sample time-series data for calculations charts (JSON).

    This endpoint returns simple arrays of dates and values for soil, climate and drought
    so the front-end can fetch realistic example series. If the DB has recent records
    this could be extended to return real time-series.
    """
    # Example time-series: last 7 days
    from datetime import datetime, timedelta

    today = datetime.utcnow().date()
    dates = [(today - timedelta(days=i)).isoformat() for i in reversed(range(7))]

    # Sample series (coherent example values)
    soil_awc = [100, 98, 96, 95, 94, 93, 92]
    climate_temp = [22.1, 22.5, 23.0, 24.2, 25.0, 24.7, 24.5]
    climate_et0 = [3.5, 3.8, 4.0, 4.1, 4.3, 4.0, 4.2]
    drought_spi1 = [-0.2, -0.3, -0.4, -0.6, -0.8, -0.9, -1.0]

    payload = {
        'dates': dates,
        'soil': {
            'labels': dates,
            'awc_mm': soil_awc,
        },
        'climate': {
            'labels': dates,
            'temperature_c': climate_temp,
            'et0_mm_day': climate_et0,
        },
        'drought': {
            'labels': dates,
            'spi1': drought_spi1,
        }
    }
    return JsonResponse(payload)


def soil_metrics_view(request):
    """Display and manage soil metrics"""
    return redirect(reverse('dashboard:metrics_browser') + '#soil')


def climate_metrics_view(request):
    """Display and manage climate metrics"""
    return redirect(reverse('dashboard:metrics_browser') + '#climate')


def drought_indices_view(request):
    """Display and manage drought indices"""
    return redirect(reverse('dashboard:metrics_browser') + '#drought')


def agricultural_metrics_view(request):
    """Display and manage agricultural metrics"""
    return redirect(reverse('dashboard:metrics_browser') + '#agricultural')


def remote_sensing_view(request):
    """Display and manage remote sensing metrics"""
    return redirect(reverse('dashboard:metrics_browser') + '#remote_sensing')


def hydrology_metrics_view(request):
    """Display and manage hydrology metrics"""
    return redirect(reverse('dashboard:metrics_browser') + '#hydrology')


# ============== DATA ENTRY VIEWS ==============

def add_soil_metrics(request):
    """Add new soil metrics"""
    if not _check_data_permission(request):
        return _deny_data_access(request)
    if request.method == 'POST':
        form = SoilMetricsForm(request.POST)
        if form.is_valid():
            region = form.cleaned_data.get('region')
            if not _check_data_permission(request, region.id if region else None):
                messages.error(request, _('You can only add data for your assigned region.'))
                return render(request, 'dashboard/metrics_form.html', {'form': form, 'title': 'Add Soil Metrics'})
            instance = form.save(commit=False)
            instance.created_by = request.user if request.user.is_authenticated else None
            instance.save()
            messages.success(request, _('Soil metrics added successfully!'))
            return redirect('dashboard:soil_metrics')
    else:
        form = SoilMetricsForm()
    
    return render(request, 'dashboard/metrics_form.html', {'form': form, 'title': 'Add Soil Metrics'})


def add_climate_metrics(request):
    """Add new climate metrics"""
    if not _check_data_permission(request):
        return _deny_data_access(request)
    if request.method == 'POST':
        form = ClimateMetricsForm(request.POST)
        if form.is_valid():
            region = form.cleaned_data.get('region')
            if not _check_data_permission(request, region.id if region else None):
                messages.error(request, _('You can only add data for your assigned region.'))
                return render(request, 'dashboard/metrics_form.html', {'form': form, 'title': 'Add Climate Metrics'})
            instance = form.save(commit=False)
            instance.created_by = request.user if request.user.is_authenticated else None
            instance.save()
            messages.success(request, _('Climate metrics added successfully!'))
            return redirect('dashboard:climate_metrics')
    else:
        form = ClimateMetricsForm()
    
    return render(request, 'dashboard/metrics_form.html', {'form': form, 'title': 'Add Climate Metrics'})


def add_drought_indices(request):
    """Add new drought indices"""
    if not _check_data_permission(request):
        return _deny_data_access(request)
    if request.method == 'POST':
        form = DroughtIndicesForm(request.POST)
        if form.is_valid():
            region = form.cleaned_data.get('region')
            if not _check_data_permission(request, region.id if region else None):
                messages.error(request, _('You can only add data for your assigned region.'))
                return render(request, 'dashboard/metrics_form.html', {'form': form, 'title': 'Add Drought Indices'})
            instance = form.save(commit=False)
            instance.created_by = request.user if request.user.is_authenticated else None
            instance.save()
            messages.success(request, _('Drought indices added successfully!'))
            return redirect('dashboard:drought_indices')
    else:
        form = DroughtIndicesForm()
    
    return render(request, 'dashboard/metrics_form.html', {'form': form, 'title': 'Add Drought Indices'})


def add_agricultural_metrics(request):
    """Add new agricultural metrics"""
    if not _check_data_permission(request):
        return _deny_data_access(request)
    if request.method == 'POST':
        form = AgriculturalMetricsForm(request.POST)
        if form.is_valid():
            region = form.cleaned_data.get('region')
            if not _check_data_permission(request, region.id if region else None):
                messages.error(request, _('You can only add data for your assigned region.'))
                return render(request, 'dashboard/metrics_form.html', {'form': form, 'title': 'Add Agricultural Metrics'})
            instance = form.save(commit=False)
            instance.created_by = request.user if request.user.is_authenticated else None
            instance.save()
            messages.success(request, _('Agricultural metrics added successfully!'))
            return redirect('dashboard:agricultural_metrics')
    else:
        form = AgriculturalMetricsForm()
    
    return render(request, 'dashboard/metrics_form.html', {'form': form, 'title': 'Add Agricultural Metrics'})


def add_remote_sensing_metrics(request):
    """Add new remote sensing metrics"""
    if not _check_data_permission(request):
        return _deny_data_access(request)
    if request.method == 'POST':
        form = RemoteSensingMetricsForm(request.POST)
        if form.is_valid():
            region = form.cleaned_data.get('region')
            if not _check_data_permission(request, region.id if region else None):
                messages.error(request, _('You can only add data for your assigned region.'))
                return render(request, 'dashboard/metrics_form.html', {'form': form, 'title': 'Add Remote Sensing Metrics'})
            instance = form.save(commit=False)
            instance.created_by = request.user if request.user.is_authenticated else None
            instance.save()
            messages.success(request, _('Remote sensing metrics added successfully!'))
            return redirect('dashboard:remote_sensing')
    else:
        form = RemoteSensingMetricsForm()
    
    return render(request, 'dashboard/metrics_form.html', {'form': form, 'title': 'Add Remote Sensing Metrics'})


def add_hydrology_metrics(request):
    """Add new hydrology metrics"""
    if not _check_data_permission(request):
        return _deny_data_access(request)
    if request.method == 'POST':
        form = HydrologyMetricsForm(request.POST)
        if form.is_valid():
            region = form.cleaned_data.get('region')
            if not _check_data_permission(request, region.id if region else None):
                messages.error(request, _('You can only add data for your assigned region.'))
                return render(request, 'dashboard/metrics_form.html', {'form': form, 'title': 'Add Hydrology Metrics'})
            instance = form.save(commit=False)
            instance.created_by = request.user if request.user.is_authenticated else None
            instance.save()
            messages.success(request, _('Hydrology metrics added successfully!'))
            return redirect('dashboard:hydrology_metrics')
    else:
        form = HydrologyMetricsForm()
    
    return render(request, 'dashboard/metrics_form.html', {'form': form, 'title': 'Add Hydrology Metrics'})


# ============== EXCEL IMPORT VIEWS ==============

def import_excel_metrics(request):
    """Import metrics from Excel file"""
    if request.method == 'POST':
        if _is_viewer_user(request.user):
            messages.error(request, _('Viewers cannot import data.'))
            return redirect('dashboard:data_ingestion')
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES['excel_file']
                metric_type = form.cleaned_data['metric_type']
                
                # Read Excel file
                df = pd.read_excel(excel_file)
                records_imported = 0
                errors = []
                
                # Process based on metric type
                result = None
                if metric_type == 'soil':
                    result = _import_soil_metrics(df, errors)
                elif metric_type == 'climate':
                    result = _import_climate_metrics(df, errors)
                elif metric_type == 'drought':
                    result = _import_drought_indices(df, errors)
                elif metric_type == 'agricultural':
                    result = _import_agricultural_metrics(df, errors)
                elif metric_type == 'remote_sensing':
                    result = _import_remote_sensing_metrics(df, errors)
                elif metric_type == 'hydrology':
                    result = _import_hydrology_metrics(df, errors)

                # Normalize result
                if isinstance(result, dict):
                    records_imported = result.get('count', 0)
                    created = result.get('created', 0)
                    updated = result.get('updated', 0)
                else:
                    records_imported = int(result or 0)
                    created = 0
                    updated = 0
                
                # Log the import
                status = 'Partial' if errors else 'Success'
                DataImportLog.objects.create(
                    source='Excel',
                    filename=excel_file.name,
                    metric_type=metric_type,
                    records_imported=records_imported,
                    status=status,
                    notes=(f"{len(errors)} errors" if errors else f"Created {created}, Updated {updated}")
                )

                messages.success(request, _('Excel import completed! {} records imported.'.format(records_imported)))
                if created or updated:
                    messages.info(request, _(f"Created: {created}, Updated: {updated}"))
                if errors:
                    messages.warning(request, _('Some records had errors: {}'.format(len(errors))))
                
                return redirect('dashboard:data_ingestion')
            except Exception as e:
                messages.error(request, _('Error importing Excel file: {}'.format(str(e))))
    else:
        form = ExcelImportForm()
    
    return render(request, 'dashboard/excel_import.html', {'form': form})


def _import_soil_metrics(df, errors):
    """Import soil metrics from DataFrame"""
    count = 0
    created = 0
    updated = 0
    for idx, row in df.iterrows():
        try:
            region = Region.objects.get(name=row.get('region', ''))
            year = ObservationYear.objects.get(label=row.get('year', ''))
            measurement_date = pd.to_datetime(row.get('measurement_date')).date() if pd.notna(row.get('measurement_date')) else None

            defaults = {
                'moisture_content_percent': float(row.get('moisture_content_percent')) if pd.notna(row.get('moisture_content_percent')) else None,
                'sand_ratio': float(row.get('sand_ratio')) if pd.notna(row.get('sand_ratio')) else None,
                'clay_ratio': float(row.get('clay_ratio')) if pd.notna(row.get('clay_ratio')) else None,
                'silt_ratio': float(row.get('silt_ratio')) if pd.notna(row.get('silt_ratio')) else None,
                'root_zone_depth_mm': int(row.get('root_zone_depth_mm')) if pd.notna(row.get('root_zone_depth_mm')) else None,
                'organic_matter_percent': float(row.get('organic_matter_percent')) if pd.notna(row.get('organic_matter_percent')) else None,
                'infiltration_rate_mmhr': float(row.get('infiltration_rate_mmhr')) if pd.notna(row.get('infiltration_rate_mmhr')) else None,
                'field_capacity_percent': float(row.get('field_capacity_percent')) if pd.notna(row.get('field_capacity_percent')) else None,
                'wilting_point_percent': float(row.get('wilting_point_percent')) if pd.notna(row.get('wilting_point_percent')) else None,
                'salinity_ece_dsm': float(row.get('salinity_ece_dsm')) if pd.notna(row.get('salinity_ece_dsm')) else None,
                'ph_level': float(row.get('ph_level')) if pd.notna(row.get('ph_level')) else None,
            }

            obj, created_flag = SoilMetrics.objects.update_or_create(
                region=region,
                year=year,
                measurement_date=measurement_date,
                defaults=defaults,
            )
            count += 1
            if created_flag:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")

    return {'count': count, 'created': created, 'updated': updated}


def _import_climate_metrics(df, errors):
    """Import climate metrics from DataFrame using batch upsert (without ON CONFLICT)."""
    from datetime import datetime
    from collections import defaultdict

    region_map = {r.name: r for r in Region.objects.all()}
    year_map = {y.label: y for y in ObservationYear.objects.all()}

    region_ids = set()
    rows_data = []  # (region, year, measurement_date, row)

    for idx, row in df.iterrows():
        try:
            region_name = row.get('region', '')
            year_label = str(row.get('year', ''))

            region = region_map.get(region_name)
            if not region:
                errors.append(f"Row {idx}: Region '{region_name}' not found")
                continue

            year = year_map.get(year_label)
            if not year:
                errors.append(f"Row {idx}: Year '{year_label}' not found")
                continue

            measurement_date = pd.to_datetime(row.get('measurement_date')).date() if pd.notna(row.get('measurement_date')) else None
            if not measurement_date:
                errors.append(f"Row {idx}: Invalid measurement_date")
                continue

            region_ids.add(region.id)
            rows_data.append((region, year, measurement_date, row))

        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")

    if not rows_data:
        return {'count': 0, 'created': 0, 'updated': 0}

    now = datetime.now()

    def _safe_val(v, caster=float):
        if pd.isna(v):
            return None
        try:
            val = caster(v)
            # Reject common sentinel values (-999, -9999, etc.)
            if isinstance(val, (int, float)) and val < -900:
                return None
            return val
        except (ValueError, TypeError):
            return None

    def _build_record(region, year, measurement_date, row):
        return ClimateMetrics(
            region=region,
            year=year,
            measurement_date=measurement_date,
            rainfall_mm=_safe_val(row.get('rainfall_mm')),
            seasonal_rainfall_variability=row.get('seasonal_rainfall_variability', ''),
            temperature_max_c=_safe_val(row.get('temperature_max_c')),
            temperature_min_c=_safe_val(row.get('temperature_min_c')),
            temperature_mean_c=_safe_val(row.get('temperature_mean_c')),
            relative_humidity_percent=_safe_val(row.get('relative_humidity_percent'), int),
            wind_speed_ms=_safe_val(row.get('wind_speed_ms')),
            solar_radiation_mjm2day=_safe_val(row.get('solar_radiation_mjm2day')),
            evapotranspiration_et0_mmday=_safe_val(row.get('evapotranspiration_et0_mmday')),
            evapotranspiration_etc_mmday=_safe_val(row.get('evapotranspiration_etc_mmday')),
            updated_at=now,
        )

    update_fields = [
        'rainfall_mm', 'seasonal_rainfall_variability',
        'temperature_max_c', 'temperature_min_c', 'temperature_mean_c',
        'relative_humidity_percent', 'wind_speed_ms',
        'solar_radiation_mjm2day', 'evapotranspiration_et0_mmday',
        'evapotranspiration_etc_mmday', 'latitude', 'longitude',
        'updated_at',
    ]

    existing = ClimateMetrics.objects.filter(
        region_id__in=region_ids,
    ).values_list('id', 'region_id', 'year_id', 'measurement_date')

    existing_keys = {}
    for pk, rid, yid, mdate in existing:
        existing_keys[(rid, yid, mdate)] = pk

    to_create = []
    to_update = []
    rainfall_by_region = defaultdict(list)

    for region, year, measurement_date, row in rows_data:
        rec = _build_record(region, year, measurement_date, row)
        key = (region.id, year.id, measurement_date)
        pk = existing_keys.get(key)
        if pk is not None:
            rec.id = pk
            to_update.append(rec)
        else:
            to_create.append(rec)

        rainfall_mm = float(row.get('rainfall_mm')) if pd.notna(row.get('rainfall_mm')) else None
        if rainfall_mm is not None:
            rainfall_by_region[region.id].append((measurement_date, rainfall_mm))

    if to_create:
        ClimateMetrics.objects.bulk_create(to_create)

    if to_update:
        ClimateMetrics.objects.bulk_update(to_update, update_fields)

    _batch_calculate_spi(dict(rainfall_by_region))

    return {
        'count': len(to_create) + len(to_update),
        'created': len(to_create),
        'updated': len(to_update),
    }


def _gamma_spi(rainfall_mm, mean, std):
    """Standardized Precipitation Index using fitted gamma distribution."""
    from scipy import stats as scipy_stats
    if std == 0 or mean == 0:
        return 0.0
    variance = std ** 2
    alpha = (mean ** 2) / variance if variance > 0 else 1.0
    scale = variance / mean if mean > 0 else 1.0
    gamma_cdf = scipy_stats.gamma.cdf(max(0, rainfall_mm), a=alpha, scale=scale)
    gamma_cdf = max(1e-15, min(1 - 1e-15, gamma_cdf))
    return float(scipy_stats.norm.ppf(gamma_cdf))


def _batch_calculate_spi(rainfall_by_region):
    """Calculate SPI for all rainfall points, grouped by region, in bulk."""
    if not rainfall_by_region:
        return

    for region_id, points in rainfall_by_region.items():
        region = Region.objects.get(id=region_id)

        all_climate = list(ClimateMetrics.objects.filter(
            region=region, rainfall_mm__isnull=False,
        ).order_by('measurement_date').values_list('measurement_date', 'rainfall_mm'))

        if len(all_climate) < 2:
            continue

        vals = [float(c[1]) for c in all_climate]
        series = pd.Series(vals)
        overall_mean = float(series.mean())
        overall_std = float(series.std())
        if overall_std == 0:
            continue

        roll_90_mean = series.rolling(90, min_periods=1).mean()
        roll_90_std = series.rolling(90, min_periods=1).std(ddof=0)
        roll_365_mean = series.rolling(365, min_periods=1).mean()
        roll_365_std = series.rolling(365, min_periods=1).std(ddof=0)

        date_to_idx = {d: i for i, (d, _) in enumerate(all_climate)}

        drought_records = []
        year_cache = {}

        for date, rainfall_mm in points:
            idx = date_to_idx.get(date)
            if idx is None:
                continue

            spi_1 = round(max(-3.29, min(3.29, _gamma_spi(rainfall_mm, overall_mean, overall_std))), 2)

            m90 = float(roll_90_mean.iloc[idx])
            s90 = float(roll_90_std.iloc[idx])
            spi_3 = round(max(-3.29, min(3.29, _gamma_spi(rainfall_mm, m90, s90))), 2) if s90 > 0 else spi_1

            m365 = float(roll_365_mean.iloc[idx])
            s365 = float(roll_365_std.iloc[idx])
            spi_12 = round(max(-3.29, min(3.29, _gamma_spi(rainfall_mm, m365, s365))), 2) if s365 > 0 else spi_1

            year_label = str(date.year)
            if year_label not in year_cache:
                year_obj, _ = ObservationYear.objects.get_or_create(label=year_label)
                year_cache[year_label] = year_obj
            year = year_cache[year_label]

            drought_records.append(DroughtIndices(
                region=region,
                year=year,
                measurement_date=date,
                spi_1month=spi_1,
                spi_3month=spi_3,
                spi_12month=spi_12,
            ))

        if drought_records:
            existing_di = DroughtIndices.objects.filter(
                region=region,
                measurement_date__in=[d.measurement_date for d in drought_records],
            ).values_list('id', 'year_id', 'measurement_date')
            di_existing = {(yid, md): pk for pk, yid, md in existing_di}

            to_create_di = []
            to_update_di = []
            for drec in drought_records:
                pk = di_existing.get((drec.year_id, drec.measurement_date))
                if pk is not None:
                    drec.id = pk
                    to_update_di.append(drec)
                else:
                    to_create_di.append(drec)

            if to_create_di:
                DroughtIndices.objects.bulk_create(to_create_di)
            if to_update_di:
                DroughtIndices.objects.bulk_update(
                    to_update_di, ['spi_1month', 'spi_3month', 'spi_12month']
                )

def _import_drought_indices(df, errors):
    """Import drought indices from DataFrame"""
    count = 0
    created = 0
    updated = 0
    for idx, row in df.iterrows():
        try:
            region = Region.objects.get(name=row.get('region', ''))
            year = ObservationYear.objects.get(label=row.get('year', ''))
            measurement_date = pd.to_datetime(row.get('measurement_date')).date() if pd.notna(row.get('measurement_date')) else None

            defaults = {
                'spi_1month': float(row.get('spi_1month')) if pd.notna(row.get('spi_1month')) else None,
                'spi_3month': float(row.get('spi_3month')) if pd.notna(row.get('spi_3month')) else None,
                'spi_12month': float(row.get('spi_12month')) if pd.notna(row.get('spi_12month')) else None,
                'spei_1month': float(row.get('spei_1month')) if pd.notna(row.get('spei_1month')) else None,
                'spei_3month': float(row.get('spei_3month')) if pd.notna(row.get('spei_3month')) else None,
                'spei_12month': float(row.get('spei_12month')) if pd.notna(row.get('spei_12month')) else None,
                'pdsi_value': float(row.get('pdsi_value')) if pd.notna(row.get('pdsi_value')) else None,
                'drought_severity_class': row.get('drought_severity_class', 'None'),
            }

            obj, created_flag = DroughtIndices.objects.update_or_create(
                region=region,
                year=year,
                measurement_date=measurement_date,
                defaults=defaults,
            )
            count += 1
            if created_flag:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")

    return {'count': count, 'created': created, 'updated': updated}


def _import_agricultural_metrics(df, errors):
    """Import agricultural metrics from DataFrame"""
    count = 0
    created = 0
    updated = 0
    for idx, row in df.iterrows():
        try:
            region = Region.objects.get(name=row.get('region', ''))
            year = ObservationYear.objects.get(label=row.get('year', ''))
            crop = CropType.objects.get(name=row.get('crop', ''))
            irrigation = None
            if pd.notna(row.get('irrigation_method')):
                try:
                    irrigation = IrrigationMethod.objects.get(name=row.get('irrigation_method'))
                except Exception:
                    irrigation = None

            measurement_date = pd.to_datetime(row.get('measurement_date')).date() if pd.notna(row.get('measurement_date')) else None

            defaults = {
                'growth_stage': row.get('growth_stage', 'Vegetative'),
                'crop_coefficient_kc': float(row.get('crop_coefficient_kc')) if pd.notna(row.get('crop_coefficient_kc')) else None,
                'crop_water_requirement_mmday': float(row.get('crop_water_requirement_mmday')) if pd.notna(row.get('crop_water_requirement_mmday')) else None,
                'yield_reduction_factor': float(row.get('yield_reduction_factor')) if pd.notna(row.get('yield_reduction_factor')) else None,
                'irrigation_method': irrigation,
                'irrigation_efficiency_percent': int(row.get('irrigation_efficiency_percent')) if pd.notna(row.get('irrigation_efficiency_percent')) else None,
                'water_applied_mm': float(row.get('water_applied_mm')) if pd.notna(row.get('water_applied_mm')) else None,
                'leaf_temperature_c': float(row.get('leaf_temperature_c')) if pd.notna(row.get('leaf_temperature_c')) else None,
                'stomatal_conductance': float(row.get('stomatal_conductance')) if pd.notna(row.get('stomatal_conductance')) else None,
            }

            obj, created_flag = AgriculturalMetrics.objects.update_or_create(
                region=region,
                year=year,
                crop=crop,
                measurement_date=measurement_date,
                defaults=defaults,
            )
            count += 1
            if created_flag:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")

    return {'count': count, 'created': created, 'updated': updated}


def _import_remote_sensing_metrics(df, errors):
    """Import remote sensing metrics from DataFrame"""
    count = 0
    created = 0
    updated = 0
    for idx, row in df.iterrows():
        try:
            region = Region.objects.get(name=row.get('region', ''))
            year = ObservationYear.objects.get(label=row.get('year', ''))
            measurement_date = pd.to_datetime(row.get('measurement_date')).date() if pd.notna(row.get('measurement_date')) else None

            defaults = {
                'ndvi': float(row.get('ndvi')) if pd.notna(row.get('ndvi')) else None,
                'ndwi': float(row.get('ndwi')) if pd.notna(row.get('ndwi')) else None,
                'land_surface_temperature_c': float(row.get('land_surface_temperature_c')) if pd.notna(row.get('land_surface_temperature_c')) else None,
                'satellite_soil_moisture_percent': float(row.get('satellite_soil_moisture_percent')) if pd.notna(row.get('satellite_soil_moisture_percent')) else None,
                'satellite_source': row.get('satellite_source', ''),
                'vegetation_condition_index': float(row.get('vegetation_condition_index')) if pd.notna(row.get('vegetation_condition_index')) else None,
                'evapotranspiration_sebal_mmday': float(row.get('evapotranspiration_sebal_mmday')) if pd.notna(row.get('evapotranspiration_sebal_mmday')) else None,
            }

            obj, created_flag = RemoteSensingMetrics.objects.update_or_create(
                region=region,
                year=year,
                measurement_date=measurement_date,
                defaults=defaults,
            )
            count += 1
            if created_flag:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")

    return {'count': count, 'created': created, 'updated': updated}


def _import_hydrology_metrics(df, errors):
    """Import hydrology metrics from DataFrame"""
    count = 0
    created = 0
    updated = 0
    for idx, row in df.iterrows():
        try:
            region = Region.objects.get(name=row.get('region', ''))
            year = ObservationYear.objects.get(label=row.get('year', ''))
            measurement_date = pd.to_datetime(row.get('measurement_date')).date() if pd.notna(row.get('measurement_date')) else None

            defaults = {
                'precipitation_mm': float(row.get('precipitation_mm')) if pd.notna(row.get('precipitation_mm')) else None,
                'evapotranspiration_mm': float(row.get('evapotranspiration_mm')) if pd.notna(row.get('evapotranspiration_mm')) else None,
                'groundwater_depth_m': float(row.get('groundwater_depth_m')) if pd.notna(row.get('groundwater_depth_m')) else None,
                'runoff_mm': float(row.get('runoff_mm')) if pd.notna(row.get('runoff_mm')) else None,
                'river_flow_m3s': float(row.get('river_flow_m3s')) if pd.notna(row.get('river_flow_m3s')) else None,
                'reservoir_storage_m3': int(row.get('reservoir_storage_m3')) if pd.notna(row.get('reservoir_storage_m3')) else None,
                'irrigation_supply_available_m3': int(row.get('irrigation_supply_available_m3')) if pd.notna(row.get('irrigation_supply_available_m3')) else None,
                'soil_water_deficit_index_mm': float(row.get('soil_water_deficit_index_mm')) if pd.notna(row.get('soil_water_deficit_index_mm')) else None,
                'water_balance_percent': float(row.get('water_balance_percent')) if pd.notna(row.get('water_balance_percent')) else None,
            }

            obj, created_flag = HydrologyMetrics.objects.update_or_create(
                region=region,
                year=year,
                measurement_date=measurement_date,
                defaults=defaults,
            )
            count += 1
            if created_flag:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")

    return {'count': count, 'created': created, 'updated': updated}


def export_metrics_excel(request, metric_type):
    """Export metrics to Excel file"""
    
    # Fetch metrics based on type
    if metric_type == 'soil':
        queryset = SoilMetrics.objects.all().values()
    elif metric_type == 'climate':
        queryset = ClimateMetrics.objects.all().values()
    elif metric_type == 'drought':
        queryset = DroughtIndices.objects.all().values()
    elif metric_type == 'agricultural':
        queryset = AgriculturalMetrics.objects.all().values()
    elif metric_type == 'remote_sensing':
        queryset = RemoteSensingMetrics.objects.all().values()
    elif metric_type == 'hydrology':
        queryset = HydrologyMetrics.objects.all().values()
    else:
        return HttpResponse('Invalid metric type', status=400)
    
    # Convert to DataFrame and export to Excel
    df = pd.DataFrame(list(queryset))
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=metric_type, index=False)
    
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="metrics_{metric_type}_{pd.Timestamp.now().strftime("%Y%m%d")}.xlsx"'
    
    return response


TEMPLATE_COLUMNS = {
    'soil': ['region', 'year', 'measurement_date', 'moisture_content_percent',
             'sand_ratio', 'clay_ratio', 'silt_ratio', 'root_zone_depth_mm',
             'organic_matter_percent', 'infiltration_rate_mmhr',
             'field_capacity_percent', 'wilting_point_percent',
             'salinity_ece_dsm', 'ph_level'],
    'climate': ['region', 'year', 'measurement_date', 'rainfall_mm',
                'seasonal_rainfall_variability', 'temperature_max_c',
                'temperature_min_c', 'temperature_mean_c',
                'relative_humidity_percent', 'wind_speed_ms',
                'solar_radiation_mjm2day', 'evapotranspiration_et0_mmday',
                'evapotranspiration_etc_mmday'],
    'drought': ['region', 'year', 'measurement_date', 'spi_1month',
                'spi_3month', 'spi_12month', 'spei_1month', 'spei_3month',
                'spei_12month', 'pdsi_value', 'drought_severity_class'],
    'agricultural': ['region', 'year', 'crop', 'measurement_date',
                     'growth_stage', 'crop_coefficient_kc',
                     'crop_water_requirement_mmday', 'yield_reduction_factor',
                     'irrigation_method', 'irrigation_efficiency_percent',
                     'water_applied_mm', 'leaf_temperature_c',
                     'stomatal_conductance'],
    'remote_sensing': ['region', 'year', 'measurement_date', 'ndvi', 'ndwi',
                       'land_surface_temperature_c',
                       'satellite_soil_moisture_percent', 'satellite_source',
                       'vegetation_condition_index',
                       'evapotranspiration_sebal_mmday'],
    'hydrology': ['region', 'year', 'measurement_date', 'precipitation_mm',
                  'evapotranspiration_mm', 'groundwater_depth_m', 'runoff_mm',
                  'river_flow_m3s', 'reservoir_storage_m3',
                  'irrigation_supply_available_m3',
                  'soil_water_deficit_index_mm', 'water_balance_percent'],
}

def download_metric_template(request, metric_type):
    """Download a blank Excel template for the given metric type."""
    if metric_type not in TEMPLATE_COLUMNS:
        return HttpResponse('Invalid metric type', status=400)

    cols = TEMPLATE_COLUMNS[metric_type]
    df = pd.DataFrame(columns=cols)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=metric_type, index=False)

    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="template_{metric_type}.xlsx"'
    return response


def import_logs_view(request):
    """View import history logs"""
    logs = DataImportLog.objects.order_by('-import_date')
    
    context = {
        'logs': logs,
        'total_records': sum(log.records_imported for log in logs),
        'success_count': logs.filter(status='Success').count(),
        'failed_count': logs.filter(status='Failed').count(),
    }
    
    return render(request, 'dashboard/import_logs.html', context)


def historical_view(request):
    """Display historical comparison and climate trends"""
    context = {
        'regions': Region.objects.all().order_by('name'),
        'years': ObservationYear.objects.all(),
        'title': _('Historical Comparison'),
    }
    return render(request, 'dashboard/historical.html', context)


def historical_export_excel(request):
    """Export historical data for a region + date range as Excel."""
    from openpyxl import Workbook
    from django.http import HttpResponse

    region_id = request.GET.get('region_id')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not all([region_id, date_from, date_to]):
        return HttpResponse('Missing required parameters', status=400)

    try:
        region = Region.objects.get(pk=region_id)
    except Region.DoesNotExist:
        return HttpResponse('Invalid region', status=404)

    from decimal import Decimal
    from .models import SoilMetrics, ClimateMetrics, DroughtIndices, RemoteSensingMetrics, HydrologyMetrics, AgriculturalMetrics

    field_map_all = {
        'soil_moisture_pct': ('SoilMetrics', 'moisture_content_percent'),
        'field_capacity_pct': ('SoilMetrics', 'field_capacity_percent'),
        'wilting_point_pct': ('SoilMetrics', 'wilting_point_percent'),
        'sand_pct': ('SoilMetrics', 'sand_ratio'),
        'clay_pct': ('SoilMetrics', 'clay_ratio'),
        'organic_matter_pct': ('SoilMetrics', 'organic_matter_percent'),
        'rainfall_mm': ('ClimateMetrics', 'rainfall_mm'),
        'temp_max_c': ('ClimateMetrics', 'temperature_max_c'),
        'temp_min_c': ('ClimateMetrics', 'temperature_min_c'),
        'temp_mean_c': ('ClimateMetrics', 'temperature_mean_c'),
        'humidity_pct': ('ClimateMetrics', 'relative_humidity_percent'),
        'wind_speed_ms': ('ClimateMetrics', 'wind_speed_ms'),
        'et0_mm': ('ClimateMetrics', 'evapotranspiration_et0_mmday'),
        'etc_mm': ('ClimateMetrics', 'evapotranspiration_etc_mmday'),
        'spi': ('DroughtIndices', 'spi_1month'),
        'spei': ('DroughtIndices', 'spei_1month'),
        'pdsi': ('DroughtIndices', 'pdsi_value'),
        'ndvi': ('RemoteSensingMetrics', 'ndvi'),
        'lst_c': ('RemoteSensingMetrics', 'land_surface_temperature_c'),
        'satellite_soil_moisture_pct': ('RemoteSensingMetrics', 'satellite_soil_moisture_percent'),
        'precipitation_mm': ('HydrologyMetrics', 'precipitation_mm'),
        'groundwater_depth_m': ('HydrologyMetrics', 'groundwater_depth_m'),
        'runoff_mm': ('HydrologyMetrics', 'runoff_mm'),
        'river_flow_m3s': ('HydrologyMetrics', 'river_flow_m3s'),
        'growth_stage': ('AgriculturalMetrics', 'growth_stage'),
        'kc': ('AgriculturalMetrics', 'crop_coefficient_kc'),
        'crop_water_requirement_mmday': ('AgriculturalMetrics', 'crop_water_requirement_mmday'),
        'yield_reduction_factor': ('AgriculturalMetrics', 'yield_reduction_factor'),
    }

    model_map = {
        'SoilMetrics': SoilMetrics,
        'ClimateMetrics': ClimateMetrics,
        'DroughtIndices': DroughtIndices,
        'RemoteSensingMetrics': RemoteSensingMetrics,
        'HydrologyMetrics': HydrologyMetrics,
        'AgriculturalMetrics': AgriculturalMetrics,
    }

    all_models_qs = {}
    for model_name, model_cls in model_map.items():
        qs = model_cls.objects.filter(
            region=region,
            measurement_date__gte=date_from,
            measurement_date__lte=date_to,
        ).order_by('measurement_date')
        all_models_qs[model_name] = list(qs)

    def _make_df(model_name, field_map_slice):
        rows = []
        for obj in all_models_qs[model_name]:
            dt = obj.measurement_date
            if dt and hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
                dt = dt.replace(tzinfo=None)
            row = {'date': dt.strftime('%Y-%m-%d') if dt else ''}
            for target_name, (src_model, src_field) in field_map_slice.items():
                val = getattr(obj, src_field, None)
                if hasattr(val, 'name'):
                    val = val.name
                elif isinstance(val, Decimal):
                    val = float(str(val))
                row[target_name] = val
            rows.append(row)
        if not rows:
            return pd.DataFrame(columns=['date'] + list(field_map_slice.keys()))
        return pd.DataFrame(rows)

    dfs = {}
    for model_name in model_map:
        slice_map = {k: v for k, v in field_map_all.items() if v[0] == model_name}
        df = _make_df(model_name, slice_map)
        if not df.empty:
            dfs[model_name] = df

    if dfs:
        from functools import reduce
        merged = reduce(lambda left, right: pd.merge(left, right, on='date', how='outer'), dfs.values())
        merged = merged.sort_values('date').reset_index(drop=True)
    else:
        merged = pd.DataFrame(columns=['date'] + list(field_map_all.keys()))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Historical_Data'

    if merged.empty:
        ws.cell(row=1, column=1, value='No data available for the selected region and date range.')
    else:
        for col_idx, col_name in enumerate(merged.columns, 1):
            ws.cell(row=1, column=col_idx, value=str(col_name))
        for row_idx, (_, row) in enumerate(merged.iterrows(), 2):
            for col_idx, col_name in enumerate(merged.columns, 1):
                val = row[col_name]
                if pd.isna(val):
                    val = ''
                ws.cell(row=row_idx, column=col_idx, value=val)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="historical_{region.name.lower()}_{date_from}_to_{date_to}.xlsx"'
    wb.save(response)
    return response


def landing_view(request):
    """Always show the login/landing page — no redirect for authenticated users."""
    return render(request, 'dashboard/login.html')


def login_view(request):
    """Display login page and handle authentication"""
    if request.user.is_authenticated:
        next_url = request.GET.get('next') or request.POST.get('next') or 'dashboard:dashboard'
        return redirect(next_url)

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            auth_login(request, user)
            next_url = request.POST.get('next') or request.GET.get('next') or 'dashboard:dashboard'
            return redirect(next_url)
        messages.error(request, _('Invalid username or password.'))

    return render(request, 'dashboard/login.html')


def logout_view(request):
    """Log out current user and redirect to login"""
    auth_logout(request)
    messages.info(request, _('You have been logged out.'))
    return redirect('dashboard:login')


@require_http_methods(["GET"])
def drought_prediction_api(request):
    """Return the latest drought prediction or generate a fresh one on demand."""
    region_id = request.GET.get('region_id')
    year_id = request.GET.get('year_id')
    refresh = request.GET.get('refresh', '0').lower() in {'1', 'true', 'yes', 'on'}
    use_llm = request.GET.get('use_llm', '0').lower() in {'1', 'true', 'yes', 'on'}

    region = None
    year = None
    if region_id:
        region = get_object_or_404(Region, pk=region_id)
    else:
        region = Region.objects.filter(name="Bizerte").first() or Region.objects.order_by('name').first()
    if year_id:
        year = get_object_or_404(ObservationYear, pk=year_id)
    else:
        year = ObservationYear.objects.order_by('-label').first()

    if region is None or year is None:
        return JsonResponse({'success': False, 'error': 'Region and year are required.'}, status=400)

    latest_prediction = DroughtPrediction.objects.filter(region=region, year=year).order_by('-generated_at').first()
    if latest_prediction is None:
        latest_prediction = DroughtPrediction.objects.filter(region=region).order_by('-generated_at').first()
    _fc_soil = SoilMetrics.objects.filter(region=region, field_capacity_percent__isnull=False).order_by('-measurement_date').first()
    _fc = float(_fc_soil.field_capacity_percent) if _fc_soil and _fc_soil.field_capacity_percent else 30.0
    if latest_prediction and not refresh:
        sm_today = float(latest_prediction.soil_moisture_today_pct or 0)
        sm_7 = float(latest_prediction.soil_moisture_7day_pct or 0)
        sm_30 = float(latest_prediction.soil_moisture_30day_pct or 0)
        return JsonResponse({
            'success': True,
            'source': 'database',
            'prediction': {
                'region': region.name,
                'year': year.label,
                'generated_at': latest_prediction.generated_at,
                'risk_scores': {
                    'today': float(latest_prediction.current_risk_score),
                    'day_7': float(latest_prediction.risk_7day),
                    'day_30': float(latest_prediction.risk_30day),
                },
                'current': {
                    'soil_moisture_pct': sm_today,
                    'soil_water_percent_of_capacity': round(min(100, (sm_today / 35) * 100), 1) if sm_today else 0,
                },
                'forecasts': {
                    'soil_moisture_7day_pct': sm_7,
                    'soil_moisture_30day_pct': sm_30,
                },
                'drivers': latest_prediction.drivers,
                'llm_explanation': latest_prediction.explanation,
            },
        })

    try:
        from .prediction_engine.pipeline import DroughtPredictionPipeline

        pipeline = DroughtPredictionPipeline()
        result = pipeline.predict_for_region(region, year, use_llm=use_llm)
        saved_prediction = pipeline.save_prediction(region, year, result)

        return JsonResponse({
            'success': True,
            'source': 'generated',
            'prediction_id': saved_prediction.id,
            'prediction': {
                'region': region.name,
                'year': year.label,
                'generated_at': saved_prediction.generated_at,
                'risk_scores': result['risk_scores'],
                'current': result['current'],
                'forecasts': result['forecasts'],
                'drivers': result['drivers'],
                'llm_explanation': result['llm_explanation'],
            },
        })
    except Exception as exc:
        if latest_prediction is not None:
            sm_today = float(latest_prediction.soil_moisture_today_pct or 0)
            return JsonResponse({
                'success': True,
                'source': 'database',
                'warning': str(exc),
                'prediction': {
                    'region': region.name,
                    'year': year.label,
                    'generated_at': latest_prediction.generated_at,
                    'risk_scores': {
                        'today': float(latest_prediction.current_risk_score),
                        'day_7': float(latest_prediction.risk_7day),
                        'day_30': float(latest_prediction.risk_30day),
                    },
                    'current': {
                        'soil_moisture_pct': sm_today,
                    'soil_water_percent_of_capacity': round(min(100, (sm_today / _fc) * 100), 1) if sm_today and _fc > 0 else 0,
                    },
                    'forecasts': {
                        'soil_moisture_7day_pct': float(latest_prediction.soil_moisture_7day_pct or 0),
                        'soil_moisture_30day_pct': float(latest_prediction.soil_moisture_30day_pct or 0),
                    },
                    'drivers': latest_prediction.drivers,
                    'llm_explanation': latest_prediction.explanation,
                },
            })
        return JsonResponse({'success': False, 'error': str(exc)}, status=500)


@require_http_methods(["GET"])
def resolve_region_api(request):
    """Given lat/lon, return the nearest region's ID and name."""
    try:
        lat = float(request.GET.get("lat", "0"))
        lon = float(request.GET.get("lon", "0"))
    except ValueError:
        return JsonResponse({"error": "Invalid coordinates"}, status=400)
    regions = Region.objects.exclude(latitude__isnull=True, longitude__isnull=True)
    nearest = None
    nearest_dist = float('inf')
    for region in regions:
        dist = _haversine_distance(lat, lon, float(region.latitude), float(region.longitude))
        if dist < nearest_dist:
            nearest_dist = dist
            nearest = region
    if nearest is None:
        first = Region.objects.order_by("name").first()
        if first:
            return JsonResponse({"region_id": first.id, "region_name": first.name, "distance_km": 0})
        return JsonResponse({"error": "No regions found"}, status=404)
    return JsonResponse({
        "region_id": nearest.id,
        "region_name": nearest.name,
        "distance_km": round(nearest_dist, 1),
    })


@require_http_methods(["POST"])
def sensor_ingest_api(request):
    """Register a new sensor pin."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)
    lat = data.get("latitude")
    lng = data.get("longitude")
    name = data.get("name", "").strip()
    sensor_type = data.get("sensor_type", "custom")
    radius = data.get("coverage_radius_km", 10)
    description = data.get("description", "")
    if not name:
        return JsonResponse({"error": "name is required"}, status=400)
    if lat is None or lng is None:
        return JsonResponse({"error": "latitude and longitude are required"}, status=400)
    # assign nearest region
    regions = Region.objects.exclude(latitude__isnull=True, longitude__isnull=True)
    nearest = None
    nearest_dist = float('inf')
    for region in regions:
        dist = _haversine_distance(lat, lng, float(region.latitude), float(region.longitude))
        if dist < nearest_dist:
            nearest_dist = dist
            nearest = region
    sensor = Sensor.objects.create(
        name=name,
        sensor_type=sensor_type,
        latitude=lat,
        longitude=lng,
        coverage_radius_km=radius,
        region=nearest,
        description=description,
        created_by=request.user if request.user.is_authenticated else None,
    )
    return JsonResponse({
        "id": sensor.id,
        "name": sensor.name,
        "sensor_type": sensor.sensor_type,
        "latitude": float(sensor.latitude),
        "longitude": float(sensor.longitude),
        "coverage_radius_km": float(sensor.coverage_radius_km),
        "region_id": sensor.region_id,
        "region_name": sensor.region.name if sensor.region else None,
    }, status=201)


@require_http_methods(["GET"])
def sensor_list_api(request):
    """Return all active sensors."""
    sensors = Sensor.objects.filter(is_active=True).select_related("region")
    data = []
    for s in sensors:
        data.append({
            "id": s.id,
            "name": s.name,
            "sensor_type": s.sensor_type,
            "latitude": float(s.latitude),
            "longitude": float(s.longitude),
            "coverage_radius_km": float(s.coverage_radius_km),
            "region_id": s.region_id,
            "region_name": s.region.name if s.region else None,
            "installed_at": s.installed_at.isoformat(),
        })
    return JsonResponse({"sensors": data})


@require_http_methods(["GET"])
def sensor_history_api(request, pk):
    """Return sensor details by ID (placeholder for future timeseries)."""
    try:
        s = Sensor.objects.get(pk=pk, is_active=True)
    except Sensor.DoesNotExist:
        return JsonResponse({"error": "Sensor not found"}, status=404)
    return JsonResponse({
        "id": s.id,
        "name": s.name,
        "sensor_type": s.sensor_type,
        "latitude": float(s.latitude),
        "longitude": float(s.longitude),
        "coverage_radius_km": float(s.coverage_radius_km),
        "region_id": s.region_id,
        "region_name": s.region.name if s.region else None,
        "installed_at": s.installed_at.isoformat(),
        "description": s.description or "",
    })


@require_http_methods(["PUT"])
def sensor_edit_api(request, pk):
    """Update an existing sensor."""
    try:
        s = Sensor.objects.get(pk=pk)
    except Sensor.DoesNotExist:
        return JsonResponse({"error": "Sensor not found"}, status=404)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)
    if "name" in data:
        name = data["name"].strip()
        if name:
            s.name = name
    if "sensor_type" in data:
        s.sensor_type = data["sensor_type"]
    if "latitude" in data and data["latitude"] is not None:
        s.latitude = data["latitude"]
    if "longitude" in data and data["longitude"] is not None:
        s.longitude = data["longitude"]
    if "coverage_radius_km" in data:
        s.coverage_radius_km = data["coverage_radius_km"]
    if "description" in data:
        s.description = data["description"]
    s.save()
    return JsonResponse({
        "id": s.id,
        "name": s.name,
        "sensor_type": s.sensor_type,
        "latitude": float(s.latitude),
        "longitude": float(s.longitude),
        "coverage_radius_km": float(s.coverage_radius_km),
        "description": s.description or "",
    })


@require_http_methods(["DELETE"])
def sensor_delete_api(request, pk):
    """Soft-delete a sensor (set is_active=False)."""
    try:
        s = Sensor.objects.get(pk=pk)
    except Sensor.DoesNotExist:
        return JsonResponse({"error": "Sensor not found"}, status=404)
    s.is_active = False
    s.save()
    return JsonResponse({"success": True})


@require_http_methods(["GET"])
def chart_data_api(request):
    """Return chart data for a given region, matching dashboard_view format."""
    region_id = request.GET.get("region_id")
    if not region_id:
        return JsonResponse({"error": "region_id required"}, status=400)
    try:
        region = Region.objects.get(pk=region_id)
    except Region.DoesNotExist:
        return JsonResponse({"error": "Region not found"}, status=404)

    now = timezone.now()
    year_ago = now - timedelta(days=365)
    year_id = request.GET.get("year_id")
    year_num = None
    if year_id:
        try:
            year_obj = ObservationYear.objects.get(pk=year_id)
            year_num = int(year_obj.label.split()[0])
        except (ObservationYear.DoesNotExist, ValueError, AttributeError):
            pass

    if year_num:
        climate_qs = ClimateMetrics.objects.filter(
            region=region, measurement_date__year=year_num
        ).order_by("-measurement_date")
        drought_qs = DroughtIndices.objects.filter(
            region=region, measurement_date__year=year_num
        ).order_by("-measurement_date")
        rs_qs = RemoteSensingMetrics.objects.filter(
            region=region, measurement_date__year=year_num
        ).order_by("-measurement_date")

        climate_rainfall = _year_series(climate_qs, "rainfall_mm")
        climate_temp = _year_series(climate_qs, "temperature_mean_c")
        drought_spi = _year_series(drought_qs, "spi_3month")
        drought_spei = _year_series(drought_qs, "spei_3month")
        ndvi_values = _year_series(rs_qs, "ndvi")

        chart_months = _year_months(climate_qs, "rainfall_mm")
    else:
        climate_qs = ClimateMetrics.objects.filter(
            region=region, measurement_date__gte=year_ago
        ).order_by("-measurement_date")
        climate_rainfall = _chart_series(climate_qs, "rainfall_mm")
        climate_temp = _chart_series(climate_qs, "temperature_mean_c")

        drought_qs = DroughtIndices.objects.filter(
            region=region, measurement_date__gte=year_ago
        ).order_by("-measurement_date")
        drought_spi = _chart_series(drought_qs, "spi_3month")
        drought_spei = _chart_series(drought_qs, "spei_3month")

        rs_qs = RemoteSensingMetrics.objects.filter(
            region=region, measurement_date__gte=year_ago
        ).order_by("-measurement_date")
        ndvi_values = _chart_series(rs_qs, "ndvi")

        chart_months = _chart_months(climate_qs, "rainfall_mm")

    if not any(chart_months):
        chart_months = ["Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
                        "Jan", "Feb", "Mar", "Apr", "May"]

    # Normalized (0-100%) versions for the overlaid climate chart
    climate_rainfall_norm = _normalize(climate_rainfall)
    climate_temp_norm = _normalize(climate_temp)
    drought_spi_norm = _normalize(drought_spi)
    ndvi_values_norm = _normalize(ndvi_values)

    spi_min, spi_max = -3, 3
    ndvi_min, ndvi_max = 0, 1

    def _labels(lo, hi):
        return [hi, hi - (hi - lo) / 3, lo + (hi - lo) / 3, lo]

    from django.core.serializers.json import DjangoJSONEncoder
    return JsonResponse({
        "climate_points_rainfall": _svg_points(climate_rainfall_norm, vmin=0, vmax=100),
        "climate_points_temp": _svg_points(climate_temp_norm, vmin=0, vmax=100),
        "climate_points_spi": _svg_points(drought_spi_norm, vmin=0, vmax=100),
        "climate_points_ndvi": _svg_points(ndvi_values_norm, vmin=0, vmax=100),
        "drought_points_spi": _svg_points(drought_spi, vmin=spi_min, vmax=spi_max),
        "drought_points_spei": _svg_points(drought_spei, vmin=spi_min, vmax=spi_max),
        "ndvi_points": _svg_points(ndvi_values, vmin=ndvi_min, vmax=ndvi_max),
        "chart_months": chart_months,
        "climate_labels": _labels(0, 100),
        "drought_labels": _labels(-3.0, 3.0),
        "ndvi_labels": _labels(0, 1.0),
        "region_name": region.name,
    })


@require_http_methods(["POST"])
def ajax_upload_file(request):
    """AJAX endpoint for drag-and-drop file upload — saves to PostgreSQL"""
    try:
        if _is_viewer_user(request.user):
            return JsonResponse({
                'success': False,
                'error': _('Viewers cannot import data.'),
            }, status=403)

        uploaded_file = request.FILES.get('file')
        metric_type = request.POST.get('metric_type', '')

        if not uploaded_file:
            return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)

        if not metric_type:
            return JsonResponse({'success': False, 'error': 'Please select a metric type'}, status=400)

        filename = uploaded_file.name.lower()

        # Read into DataFrame
        if filename.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
        elif filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(uploaded_file)
        else:
            return JsonResponse({
                'success': False,
                'error': 'Unsupported file format. Use .csv, .xlsx, or .xls'
            }, status=400)

        if df.empty:
            return JsonResponse({'success': False, 'error': 'The uploaded file contains no data'}, status=400)

        records_imported = 0
        errors = []
        created = 0
        updated = 0

        # Use existing import functions
        result = None
        if metric_type == 'soil':
            result = _import_soil_metrics(df, errors)
        elif metric_type == 'climate':
            result = _import_climate_metrics(df, errors)
        elif metric_type == 'drought':
            result = _import_drought_indices(df, errors)
        elif metric_type == 'agricultural':
            result = _import_agricultural_metrics(df, errors)
        elif metric_type == 'remote_sensing':
            result = _import_remote_sensing_metrics(df, errors)
        elif metric_type == 'hydrology':
            result = _import_hydrology_metrics(df, errors)
        else:
            return JsonResponse({'success': False, 'error': f'Unknown metric type: {metric_type}'}, status=400)

        # Normalize result which may be a dict with counts
        if isinstance(result, dict):
            records_imported = result.get('count', 0)
            created = result.get('created', 0)
            updated = result.get('updated', 0)
        else:
            # backward compatibility: integer
            records_imported = int(result or 0)

        # Log the import
        status = 'Failed' if records_imported == 0 and errors else ('Partial' if errors else 'Success')
        username = request.user.username if request.user.is_authenticated else 'Anonymous'

        DataImportLog.objects.create(
            source='Excel',
            filename=uploaded_file.name,
            metric_type=metric_type,
            records_imported=records_imported,
            imported_by=username,
            status=status,
            notes=f"{len(errors)} errors" if errors else "No errors",
            error_details=json.dumps(errors[:20]) if errors else None,
        )

        return JsonResponse({
            'success': True,
            'message': f'Import complete: {records_imported} records saved',
            'records_imported': records_imported,
            'total_rows': len(df),
            'error_count': len(errors),
            'errors': errors[:5],  # Return first 5 errors for display
            'created': created,
            'updated': updated,
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Upload failed: {str(e)}'
        }, status=500)


@login_required(login_url='dashboard:login')
def admin_panel(request):
    """Admin user management panel"""
    if not _is_admin_user(request.user):
        return _deny_admin_access(request)

    users = User.objects.all().select_related('profile').order_by('-date_joined')

    if request.method == 'POST':
        form = UserCreateForm(request.POST, user=request.user)
        if form.is_valid():
            user = form.save()
            messages.success(request, _('User "{}" created successfully!').format(user.username))
            return redirect('dashboard:admin_panel')
    else:
        form = UserCreateForm(user=request.user)

    context = {
        'form': form,
        'users': users,
    }
    return render(request, 'dashboard/admin_panel.html', context)


@login_required(login_url='dashboard:login')
def admin_edit_user(request, user_id):
    """Edit user role and region assignment"""
    if not _is_admin_user(request.user):
        return _deny_admin_access(request)

    target_user = get_object_or_404(User, pk=user_id)
    profile = target_user.profile

    if request.method == 'POST':
        role = request.POST.get('role')
        region_id = request.POST.get('region')

        if role in dict(UserProfile.ROLE_CHOICES):
            profile.role = role
        if region_id:
            profile.region_id = region_id
        else:
            profile.region = None
        profile.save()

        messages.success(request, _('User profile updated!'))
        return redirect('dashboard:admin_panel')

    context = {
        'target_user': target_user,
        'regions': Region.objects.all(),
        'role_choices': UserProfile.ROLE_CHOICES,
    }
    return render(request, 'dashboard/admin_edit_user.html', context)


@login_required(login_url='dashboard:login')
@require_http_methods(["POST"])
def admin_delete_user(request, user_id):
    """Delete a user (superadmin only)."""
    if not _is_admin_user(request.user):
        return _deny_admin_access(request)

    target_user = get_object_or_404(User, pk=user_id)
    if target_user == request.user:
        messages.error(request, _('You cannot delete your own account.'))
        return redirect('dashboard:admin_panel')

    target_user.delete()
    messages.success(request, _('User deleted successfully.'))
    return redirect('dashboard:admin_panel')
